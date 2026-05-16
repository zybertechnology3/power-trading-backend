"""
Telemetry API Endpoints
Handles power system operational data (generators, transmission lines, busbar voltage)
Supports bulk import from Excel and time series queries
"""

from fastapi import APIRouter, HTTPException, Query, UploadFile, File
from datetime import datetime, timezone
from typing import Optional
from bson.errors import InvalidId
from bson.objectid import ObjectId
import io
from openpyxl import load_workbook

from app.db.database import get_db
from app.schemas.telemetry import (
    TelemetryCreate,
    TelemetryResponse,
    TelemetryListResponse,
    BulkImportResponse,
    TelemetryStatsResponse,
    UnitsLoad,
    TransmissionLines,
)

router = APIRouter(prefix="/telemetry", tags=["telemetry"])


# ===== HELPER FUNCTIONS =====

def parse_excel_telemetry(file_content: bytes) -> list[TelemetryCreate]:
    """
    Parse telemetry data from Excel file
    Expected columns: TIME, DATE, UNITS_LOAD (UNIT 1-4), 
    TRANSMISSION_LINES (LINE A-D), TOTAL LHPC EXPORT, BUSBAR (kV)
    """
    records = []
    
    try:
        # Load Excel workbook from bytes
        wb = load_workbook(io.BytesIO(file_content))
        ws = wb.active
        
        # Skip header row (row 1)
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]:  # Skip empty rows
                break
            
            # Skip MAX/TOTAL rows indicated in the data
            if isinstance(row[0], str) and ("MAX" in row[0].upper() or "TOTAL" in row[0].upper()):
                continue
            
            try:
                # Parse time and date
                time_str = str(row[0]).strip()  # "1:00", "2:00", etc.
                date_obj = row[12]  # DATE column
                
                if not isinstance(date_obj, datetime):
                    continue
                
                # Combine date with hour
                hour = int(time_str.split(":")[0])
                timestamp = date_obj.replace(hour=hour, minute=0, second=0)
                
                # Parse units load data (columns 1-5)
                units_load = UnitsLoad(
                    unit_1=float(row[1]),  # UNIT 1
                    unit_2=float(row[2]),  # UNIT 2
                    unit_3=float(row[3]),  # UNIT 3
                    unit_4=float(row[4]),  # UNIT 4
                    total=float(row[5])     # TOTAL
                )
                
                # Parse transmission lines (columns 6-10)
                transmission_lines = TransmissionLines(
                    line_c=float(row[6]),           # LINE C
                    line_d=float(row[7]),           # LINE D
                    line_a=float(row[8]),           # LINE A
                    line_b=float(row[9]),           # LINE B
                    total_export=float(row[10])     # TOTAL LHPC EXPORT
                )
                
                # Parse busbar voltage (column 11)
                busbar_voltage_kv = float(row[11])
                
                # Create telemetry record
                record = TelemetryCreate(
                    timestamp=timestamp,
                    units_load=units_load,
                    transmission_lines=transmission_lines,
                    busbar_voltage_kv=busbar_voltage_kv
                )
                records.append(record)
                
            except (ValueError, IndexError, TypeError):
                # Skip malformed rows
                continue
        
        return records
    
    except Exception as e:
        raise ValueError(f"Failed to parse Excel file: {str(e)}")


# ===== TELEMETRY ENDPOINTS =====

@router.post("/bulk-import", response_model=BulkImportResponse)
async def bulk_import_telemetry(file: UploadFile = File(...)):
    """
    Bulk import telemetry data from Excel file
    Expects standard hourly format with operational measurements
    """
    db = get_db()
    telemetry_collection = db["power_system_telemetry"]
    
    try:
        # Read file content
        file_content = await file.read()
        
        # Parse Excel data
        records = parse_excel_telemetry(file_content)
        
        if not records:
            raise HTTPException(
                status_code=400,
                detail="No valid records found in Excel file"
            )
        
        # Prepare documents for bulk insert
        documents = []
        for record in records:
            now = datetime.now(timezone.utc)
            doc = {
                "timestamp": record.timestamp,
                "metadata": {
                    "date": record.timestamp.date().isoformat(),
                },
                "units_load": record.units_load.model_dump(),
                "transmission_lines": record.transmission_lines.model_dump(),
                "busbar_voltage_kv": record.busbar_voltage_kv,
                "created_at": now,
                "updated_at": now
            }
            documents.append(doc)
        
        # Insert all documents
        result = telemetry_collection.insert_many(documents)
        
        return BulkImportResponse(
            imported=len(result.inserted_ids),
            failed=0,
            errors=[]
        )
    
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(
            status_code=500,
            detail=f"Import failed: {str(e)}"
        )


@router.get("", response_model=TelemetryListResponse)
def list_telemetry(
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    start_time: Optional[datetime] = Query(None),
    end_time: Optional[datetime] = Query(None),
):
    """
    List telemetry records with optional time range filtering
    """
    db = get_db()
    telemetry_collection = db["power_system_telemetry"]
    
    # Build query filter
    query_filter = {}
    
    if start_time or end_time:
        query_filter["timestamp"] = {}
        if start_time:
            query_filter["timestamp"]["$gte"] = start_time
        if end_time:
            query_filter["timestamp"]["$lte"] = end_time
    
    # Get total count
    total = telemetry_collection.count_documents(query_filter)
    
    # Execute query with pagination (most recent first)
    records = list(
        telemetry_collection.find(query_filter)
        .sort("timestamp", -1)
        .skip(skip)
        .limit(limit)
    )
    
    # Convert ObjectIds to strings
    for record in records:
        record["_id"] = str(record["_id"])
    
    page = (skip // limit) + 1 if limit > 0 else 1
    
    return TelemetryListResponse(
        records=records,
        total=total,
        page=page,
        page_size=limit
    )


@router.get("/{telemetry_id}", response_model=TelemetryResponse)
def get_telemetry(telemetry_id: str):
    """Get a specific telemetry record by ID"""
    db = get_db()
    telemetry_collection = db["power_system_telemetry"]
    
    try:
        record_oid = ObjectId(telemetry_id)
    except (InvalidId, TypeError):
        raise HTTPException(status_code=400, detail="Invalid telemetry ID format")
    
    record = telemetry_collection.find_one({"_id": record_oid})
    if not record:
        raise HTTPException(status_code=404, detail="Telemetry record not found")
    
    record["_id"] = str(record["_id"])
    return record


@router.get("/stats/summary", response_model=TelemetryStatsResponse)
def get_telemetry_stats(
    start_time: Optional[datetime] = Query(None),
    end_time: Optional[datetime] = Query(None),
):
    """
    Get statistics for telemetry data
    Includes averages and extremes for key measurements
    """
    db = get_db()
    telemetry_collection = db["power_system_telemetry"]
    
    # Build query filter
    query_filter = {}
    if start_time or end_time:
        query_filter["timestamp"] = {}
        if start_time:
            query_filter["timestamp"]["$gte"] = start_time
        if end_time:
            query_filter["timestamp"]["$lte"] = end_time
    
    # Get all records (could optimize with aggregation pipeline for large datasets)
    records = list(telemetry_collection.find(query_filter))
    
    if not records:
        raise HTTPException(status_code=404, detail="No telemetry records found")
    
    # Calculate statistics
    total_records = len(records)
    
    # Extract timestamps for date range
    timestamps = [r["timestamp"] for r in records]
    date_range_start = min(timestamps)
    date_range_end = max(timestamps)
    
    # Calculate averages and extremes
    avg_total_generation = sum(r["units_load"]["total"] for r in records) / total_records
    avg_total_transmission = sum(r["transmission_lines"]["total_export"] for r in records) / total_records
    
    busbar_voltages = [r["busbar_voltage_kv"] for r in records]
    avg_busbar_voltage = sum(busbar_voltages) / total_records
    max_busbar_voltage = max(busbar_voltages)
    min_busbar_voltage = min(busbar_voltages)
    
    return TelemetryStatsResponse(
        total_records=total_records,
        date_range_start=date_range_start,
        date_range_end=date_range_end,
        avg_total_generation_mw=avg_total_generation,
        avg_total_transmission_mw=avg_total_transmission,
        avg_busbar_voltage_kv=avg_busbar_voltage,
        max_busbar_voltage_kv=max_busbar_voltage,
        min_busbar_voltage_kv=min_busbar_voltage
    )
