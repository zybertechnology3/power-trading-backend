import os
import re
import shutil
import subprocess
import tempfile
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Callable, Literal, Optional

import openpyxl
from dotenv import load_dotenv
from pymongo import UpdateOne
from selenium import webdriver
from selenium.common.exceptions import (
    ElementClickInterceptedException,
    StaleElementReferenceException,
    TimeoutException,
    WebDriverException,
)
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.firefox import GeckoDriverManager
from app.core.time_of_use import get_time_of_use_period

BASE_URL = "https://trading.sappmtp.com"
LOGIN_URL = f"{BASE_URL}/account/login?returnUrl=%2F"
INBOX_URL = f"{BASE_URL}/mdd/message-inbox"
AREA_RESULTS_TEST_URL = f"{BASE_URL}/amt/prices-and-turnover-X-dam"
DOWNLOAD_DIR = Path(__file__).resolve().parent / "downloads"
CONSTRAINED_AREA_SUBJECT_TEMPLATE = "MTP - DAM - Constrained Area Results for {delivery_date}"
CONSTRAINED_AREA_DATA_SOURCE = "SAPP_MTP_DAM_CONSTRAINED_AREA_RESULTS"
UNCONSTRAINED_AREA_SUBJECT_TEMPLATE = (
    "MTP - DAM - Unconstrained Results for {delivery_date}"
)
UNCONSTRAINED_AREA_DATA_SOURCE = "SAPP_MTP_DAM_UNCONSTRAINED_RESULTS"
PARTICIPANT_PORTFOLIO_SUBJECT_TEMPLATE = (
    "MTP - DAM - Participant Portfolio Results for {delivery_date}"
)
PARTICIPANT_PORTFOLIO_DATA_SOURCE = "SAPP_MTP_DAM_PARTICIPANT_PORTFOLIO_RESULTS"
PARTICIPANT_PORTFOLIO_FPM_W_SUBJECT_TEMPLATE = (
    "MTP - FPM-W - Participant Portfolio Results for {delivery_date}"
)
PARTICIPANT_PORTFOLIO_FPM_W_DATA_SOURCE = "SAPP_MTP_FPM_W_PARTICIPANT_PORTFOLIO_RESULTS"
PARTICIPANT_PORTFOLIO_FPM_M_SUBJECT_TEMPLATE = (
    "MTP - FPM-M - Participant Portfolio Results for {delivery_date}"
)
PARTICIPANT_PORTFOLIO_FPM_M_DATA_SOURCE = "SAPP_MTP_FPM_M_PARTICIPANT_PORTFOLIO_RESULTS"
TRADING_INVOICE_SUBJECT_TEMPLATE = (
    "MTP - Trading Invoice / Credit Note for {delivery_date}"
)
TRADING_INVOICE_DATA_SOURCE = "SAPP_MTP_TRADING_INVOICE_CREDIT_NOTE"
TRADING_INVOICE_HOURLY_DATA_SOURCE = "SAPP_MTP_TRADING_INVOICE_HOURLY_DETAIL"
TRADING_INVOICE_HOURLY_COLLECTION = "sapp_trading_invoice_hourly_details"
INTERNAL_COLLECTION_FIELD = "_collection_name"
INTERNAL_UNIQUE_KEY_FIELDS_FIELD = "_unique_key_fields"
INTERNAL_UNSET_FIELDS_FIELD = "_unset_fields"
MAX_INBOX_PAGES_TO_SEARCH = 75
INBOX_GRID_READY_TIMEOUT = 15
INBOX_PAGE_CHANGE_TIMEOUT = 30
MESSAGE_CLICK_TIMEOUT = 3
NEXT_PAGE_SELECTOR = (
    "button.k-pager-nav[title='Go to the next page'], "
    "button.k-pager-nav[aria-label='Go to the next page']"
)
PAGE_NUMBER_INPUT_SELECTOR = (
    "input[role='spinbutton'][title='Page Number'], "
    "input[aria-label='Type a page number'], "
    ".k-pager input.k-input-inner, "
    ".k-pager input[role='spinbutton']"
)
INBOX_GRID_SELECTOR = ".k-grid, [role='grid']"
INBOX_LOADING_SELECTOR = ".k-loading-mask, .k-i-loading, .k-loading-image"

# stress test for all possible scenarios:

PortfolioMarket = Literal["dam", "fpm_w", "fpm_m"]


@dataclass(frozen=True)
class SappExtractionJob:
    """
    Configuration for one SAPP inbox document type.

    Add new document types by creating another job with a subject template,
    parser, target Mongo collection, and unique-key fields.
    """

    name: str
    subject_template: str
    data_source: str
    collection_name: str
    unique_key_fields: tuple[str, ...]
    extractor: Callable[[Path, "SappExtractionJob"], list[dict]]
    attachment_extension: str = ".xlsx"

    def build_subject(self, delivery_date: date) -> str:
        return self.subject_template.format(
            delivery_date=delivery_date.strftime("%Y/%m/%d")
        )


def build_message_subject(delivery_date: date, subject_template: str) -> str:
    return subject_template.format(delivery_date=delivery_date.strftime("%Y/%m/%d"))


def normalize_subject_text(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", value.lower()).strip()


def delivery_date_subject_variants(delivery_date: date) -> set[str]:
    return {
        delivery_date.strftime("%Y/%m/%d"),
        delivery_date.strftime("%Y-%m-%d"),
        delivery_date.strftime("%d/%m/%Y"),
        delivery_date.strftime("%d-%m-%Y"),
    }


def job_subject_keywords(job: SappExtractionJob) -> list[str]:
    prefix = job.subject_template.split("{delivery_date}")[0]
    normalized_prefix = normalize_subject_text(prefix)
    stopwords = {"for", "the", "and"}
    return [
        token
        for token in normalized_prefix.split()
        if token and token not in stopwords
    ]


def subject_matches_job_and_date(
    subject: str,
    job: SappExtractionJob,
    delivery_date: date,
) -> bool:
    normalized_subject = normalize_subject_text(subject)
    if not all(keyword in normalized_subject for keyword in job_subject_keywords(job)):
        return False

    subject_date_tokens = {
        token
        for token in delivery_date_subject_variants(delivery_date)
        if normalize_subject_text(token) in normalized_subject
    }
    return bool(subject_date_tokens)


def parse_delivery_date_value(value):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
        try:
            return datetime.fromisoformat(value).date()
        except ValueError:
            return None
    return None


def parse_hour(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        hour = int(value)
        return hour if 1 <= hour <= 24 else None

    text = str(value).strip()
    interval_match = re.fullmatch(r"(\d{1,2})\s*-\s*(\d{1,2})", text)
    if interval_match:
        start_hour = int(interval_match.group(1))
        return start_hour + 1 if 0 <= start_hour <= 23 else None

    match = re.search(r"\d+", text)
    if not match:
        return None
    hour = int(match.group(0))
    return hour if 1 <= hour <= 24 else None


def to_float(value) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def normalize_excel_header(value) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip()).lower()


def find_labeled_date_in_sheet(sheet, labels: list[str]) -> Optional[date]:
    for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
        if not row:
            continue
        for idx, value in enumerate(row):
            if isinstance(value, str) and value.strip() in labels:
                candidate = row[idx + 1] if idx + 1 < len(row) else None
                return parse_delivery_date_value(candidate)
    return None


def find_delivery_date_in_sheet(sheet) -> Optional[date]:
    return find_labeled_date_in_sheet(sheet, ["Delivery Date:"])


def find_portfolio_reference_date_in_sheet(sheet) -> Optional[date]:
    return find_labeled_date_in_sheet(
        sheet,
        [
            "Delivery Date:",
            "Delivery Week:",
            "Delivery Month:",
        ],
    )


def next_month_start(reference_date: date) -> date:
    if reference_date.month == 12:
        return date(reference_date.year + 1, 1, 1)
    return date(reference_date.year, reference_date.month + 1, 1)


def month_end(month_start_date: date) -> date:
    return next_month_start(month_start_date) - timedelta(days=1)


def hour_label_from_hour(hour: int) -> str:
    return f"{hour - 1:02d}-{hour:02d}" if hour < 24 else "23-24"


def normalize_portfolio_product(value) -> Optional[str]:
    normalized = normalize_excel_header(value)
    aliases = {
        "off-peak": "off_peak",
        "off peak": "off_peak",
        "off_peak": "off_peak",
        "peak": "peak",
        "standard": "standard",
    }
    return aliases.get(normalized)


def infer_portfolio_market(sheet, header_row_index: int, fallback: PortfolioMarket) -> PortfolioMarket:
    for row_idx in range(max(1, header_row_index - 6), header_row_index + 1):
        row = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, sheet.max_column + 1)]
        labels = " | ".join(
            normalize_excel_header(value)
            for value in row
            if value is not None and str(value).strip()
        )
        if "fpm-w" in labels or "mtp - fpm-w" in labels or "mtp fpm-w" in labels:
            return "fpm_w"
        if "fpm-m" in labels or "mtp - fpm-m" in labels or "mtp fpm-m" in labels:
            return "fpm_m"
        if "dam" in labels or "mtp - dam" in labels or "mtp dam" in labels:
            return "dam"
    return fallback


def period_range_for_portfolio_market(
    market: PortfolioMarket,
    reference_date: date,
) -> tuple[date, date]:
    if market == "dam":
        return reference_date, reference_date
    if market == "fpm_w":
        period_start_date = reference_date
        return period_start_date, period_start_date + timedelta(days=6)

    period_start_date = date(reference_date.year, reference_date.month, 1)
    return period_start_date, month_end(period_start_date)


def safe_divide(numerator: Optional[float], denominator: Optional[float]) -> Optional[float]:
    if numerator is None or denominator in (None, 0):
        return None
    return numerator / denominator


def get_label_value(sheet, label: str):
    for row in sheet.iter_rows(values_only=True):
        for idx, value in enumerate(row):
            if isinstance(value, str) and value.strip() == label:
                return row[idx + 1] if idx + 1 < len(row) else None
    return None


def find_summary_value_columns(sheet) -> tuple[int, int, Optional[int]]:
    for row in sheet.iter_rows(values_only=True):
        normalized = [normalize_excel_header(value) for value in row]
        if "mwh" in normalized and "total" in normalized:
            mwh_col = normalized.index("mwh")
            total_col = normalized.index("total")
            price_col = (
                normalized.index("usd/mwh")
                if "usd/mwh" in normalized
                else None
            )
            return mwh_col, total_col, price_col
    raise RuntimeError("Could not find invoice summary value columns.")


def get_summary_row_values(
    sheet,
    label: str,
    mwh_col: int,
    total_col: int,
    price_col: Optional[int],
) -> dict:
    for row in sheet.iter_rows(values_only=True):
        if not row:
            continue
        first_cell = row[0] if len(row) > 0 else None
        if not isinstance(first_cell, str) or first_cell.strip() != label:
            continue

        mwh = to_float(row[mwh_col] if mwh_col < len(row) else None)
        amount = to_float(row[total_col] if total_col < len(row) else None)
        workbook_average_price = (
            to_float(row[price_col])
            if price_col is not None and price_col < len(row)
            else None
        )
        return {
            "mwh": mwh,
            "amount_usd": amount,
            "average_price_usd_per_mwh": (
                workbook_average_price
                if workbook_average_price is not None
                else safe_divide(amount, mwh)
            ),
        }

    return {
        "mwh": None,
        "amount_usd": None,
        "average_price_usd_per_mwh": None,
    }


def build_invoice_market_section(
    sheet,
    market_label: str,
    mwh_col: int,
    total_col: int,
    price_col: Optional[int],
) -> dict:
    purchases = get_summary_row_values(
        sheet,
        f"{market_label} Total Purchases",
        mwh_col,
        total_col,
        price_col,
    )
    sales = get_summary_row_values(
        sheet,
        f"{market_label} Total Sales",
        mwh_col,
        total_col,
        price_col,
    )

    purchase_mwh = purchases["mwh"] or 0.0
    purchase_amount = purchases["amount_usd"] or 0.0
    sales_mwh = sales["mwh"] or 0.0
    sales_amount = sales["amount_usd"] or 0.0
    net_mwh = purchase_mwh - sales_mwh
    net_amount = purchase_amount + sales_amount

    has_purchases = purchase_mwh != 0.0 or purchase_amount != 0.0
    has_sales = sales_mwh != 0.0 or sales_amount != 0.0

    if has_purchases and (
        not has_sales or abs(purchase_amount) >= abs(sales_amount)
    ):
        direction = "PURCHASE"
        display_mwh = purchase_mwh
        display_amount = purchase_amount
    elif has_sales:
        direction = "SALE"
        display_mwh = sales_mwh
        display_amount = sales_amount
    else:
        direction = "NONE"
        display_mwh = max(purchase_mwh, sales_mwh)
        display_amount = max(purchase_amount, sales_amount)

    return {
        "direction": direction,
        "mwh": display_mwh,
        "amount_usd": display_amount,
        "average_price_usd_per_mwh": safe_divide(display_amount, display_mwh),
        "purchases": purchases,
        "sales": sales,
        "net_mwh": net_mwh,
        "net_amount_usd": net_amount,
        "net_average_price_usd_per_mwh": safe_divide(abs(net_amount), abs(net_mwh)),
    }


def has_invoice_section_activity(section: dict) -> bool:
    return any(
        (section.get(group, {}).get(field) or 0.0) != 0.0
        for group in ("purchases", "sales")
        for field in ("mwh", "amount_usd")
    )


def find_detail_header_columns(sheet) -> Optional[dict[str, int]]:
    header_row_index = None
    normalized_headers = None
    rows = list(sheet.iter_rows(values_only=True))
    for row_idx, row in enumerate(rows, start=1):
        normalized = [normalize_excel_header(value) for value in row]
        if (
            "traded purchases" in normalized
            or "traded sales" in normalized
            or "purchase turnover" in normalized
        ):
            header_row_index = row_idx
            normalized_headers = normalized
            break

    if header_row_index is None or normalized_headers is None:
        return None

    hour_col = (
        normalized_headers.index("hour")
        if "hour" in normalized_headers
        else None
    )
    if hour_col is None and header_row_index < len(rows):
        next_row = rows[header_row_index]
        normalized_next_row = [normalize_excel_header(value) for value in next_row]
        if "hour" in normalized_next_row:
            hour_col = normalized_next_row.index("hour")

    def find_col(header: str) -> Optional[int]:
        return (
            normalized_headers.index(header)
            if header in normalized_headers
            else None
        )

    wheeling_col = None
    for col_idx, header in enumerate(normalized_headers):
        if (
            header.endswith(" wheeling cost")
            and "total wheeling cost" not in header
            and "apportionment" not in header
        ):
            wheeling_col = col_idx
            break

    return {
        "header_row": header_row_index,
        "hour": hour_col,
        "price": find_col("price"),
        "traded_purchases_mwh": find_col("traded purchases"),
        "traded_sales_mwh": find_col("traded sales"),
        "purchase_turnover_usd": find_col("purchase turnover"),
        "sale_turnover_usd": find_col("sale turnover"),
        "admin_fees_usd": find_col("admin fees"),
        "wheeling_cost_usd": wheeling_col,
    }


def extract_trading_invoice_hourly_details(
    workbook,
    delivery_date: date,
    file_path: Path,
) -> list[dict]:
    detail_records = []
    for sheet in workbook.worksheets:
        if "details" not in sheet.title.lower():
            continue

        columns = find_detail_header_columns(sheet)
        if not columns or columns["hour"] is None:
            continue

        market = (
            sheet.title.replace(" Details", "")
            .strip()
            .lower()
            .replace("-", "_")
            .replace(" ", "_")
        )
        for row in sheet.iter_rows(
            min_row=columns["header_row"] + 1,
            values_only=True,
        ):
            hour_col = columns["hour"]
            if not row or hour_col >= len(row) or row[hour_col] is None:
                continue

            hour = parse_hour(row[hour_col])
            if hour is None:
                continue

            def get_number(field_name: str) -> Optional[float]:
                col_idx = columns[field_name]
                if col_idx is None or col_idx >= len(row):
                    return None
                return to_float(row[col_idx])

            values = {
                "price_usd_per_mwh": get_number("price"),
                "traded_purchases_mwh": get_number("traded_purchases_mwh"),
                "traded_sales_mwh": get_number("traded_sales_mwh"),
                "purchase_turnover_usd": get_number("purchase_turnover_usd"),
                "sale_turnover_usd": get_number("sale_turnover_usd"),
                "admin_fees_usd": get_number("admin_fees_usd"),
                "wheeling_cost_usd": get_number("wheeling_cost_usd"),
            }
            if all(value is None for value in values.values()):
                continue

            detail_records.append(
                {
                    INTERNAL_COLLECTION_FIELD: TRADING_INVOICE_HOURLY_COLLECTION,
                    INTERNAL_UNIQUE_KEY_FIELDS_FIELD: (
                        "delivery_date",
                        "market",
                        "hour",
                    ),
                    "timestamp": delivery_timestamp(delivery_date, hour),
                    "delivery_date": delivery_date.isoformat(),
                    "market": market,
                    "hour": hour,
                    "hour_label": str(row[hour_col]).strip(),
                    **values,
                    "metadata": {
                        "data_source": TRADING_INVOICE_HOURLY_DATA_SOURCE,
                        "source_file": file_path.name,
                        "market": market,
                    },
                    "source_file": file_path.name,
                }
            )

    return detail_records


def delivery_timestamp(delivery_date: date, hour: int) -> datetime:
    return datetime.combine(delivery_date, datetime.min.time()) + timedelta(hours=hour - 1)


def load_config():
    env_path = Path(__file__).resolve().parent / ".env"
    if env_path.exists():
        load_dotenv(dotenv_path=env_path)

    username = os.getenv("SAPP_USERNAME")
    password = os.getenv("SAPP_PASSWORD")
    if not username or not password:
        raise ValueError("Please set SAPP_USERNAME and SAPP_PASSWORD as environment variables.")

    print("[1/7] ✅ Loaded credentials from .env")
    return username, password


def create_driver(download_dir: Path, headless: bool = False):
    download_dir.mkdir(parents=True, exist_ok=True)
    print(f"[2/7] 🔧 Creating Firefox driver and setting download folder to: {download_dir}")

    options = Options()
    options.add_argument("--width=1366")
    options.add_argument("--height=900")
    if headless:
        options.add_argument("-headless")
        options.headless = True
    options.set_preference("browser.download.folderList", 2)
    options.set_preference("browser.download.dir", str(download_dir.resolve()))
    options.set_preference("browser.download.useDownloadDir", True)
    options.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet,application/octet-stream")
    options.set_preference("browser.helperApps.alwaysAsk.force", False)
    options.set_preference("browser.download.manager.showWhenStarting", False)
    options.set_preference("pdfjs.disabled", True)
    options.set_preference("browser.download.manager.showWhenStarting", False)
    options.set_preference("browser.download.manager.focusWhenStarting", False)
    options.set_preference("browser.shell.checkDefaultBrowser", False)

    if headless:
        os.environ["MOZ_HEADLESS"] = "1"
    else:
        os.environ.pop("MOZ_HEADLESS", None)
    os.environ["MOZ_DISABLE_CONTENT_SANDBOX"] = "1"

    driver_path = shutil.which("geckodriver") or GeckoDriverManager().install()
    geckodriver_log_path = Path(tempfile.gettempdir()) / "geckodriver.log"
    print(f"[2/7] Using geckodriver at: {driver_path}")
    print_subprocess_version(["firefox", "--version"])
    print_subprocess_version([driver_path, "--version"])

    service = Service(executable_path=driver_path, log_output=str(geckodriver_log_path))
    try:
        driver = webdriver.Firefox(service=service, options=options)
    except WebDriverException:
        print_geckodriver_log(geckodriver_log_path)
        raise
    driver.set_window_size(1366, 900)
    driver.implicitly_wait(0)
    return driver


def print_subprocess_version(command: list[str]):
    try:
        result = subprocess.run(
            command,
            check=False,
            capture_output=True,
            text=True,
            timeout=10,
        )
        output = (result.stdout or result.stderr).strip()
        print(f"[2/7] {' '.join(command)}: {output}")
    except Exception as exc:
        print(f"[2/7] Could not run {' '.join(command)}: {exc}")


def print_geckodriver_log(log_path: Path):
    try:
        if log_path.exists():
            print("[2/7] Geckodriver log:")
            print(log_path.read_text(errors="replace")[-4000:])
        else:
            print(f"[2/7] Geckodriver log was not created at {log_path}")
    except Exception as exc:
        print(f"[2/7] Could not read geckodriver log: {exc}")


def login(driver, username: str, password: str):
    print("[3/7] 🔐 Navigating to login page")
    driver.get(LOGIN_URL)
    # TODO: update selectors according to the actual login page fields
    username_selector = "input[id='login-input-user-name-or-email-address']"
    password_selector = "input[id='password']"
    submit_selector = "button[type='submit']"

    WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, username_selector)))

    driver.find_element(By.CSS_SELECTOR, username_selector).send_keys(username)
    driver.find_element(By.CSS_SELECTOR, password_selector).send_keys(password)
    driver.find_element(By.CSS_SELECTOR, submit_selector).click()

    print("[3/7] ⏳ Submitted login form, waiting for redirect")
    time.sleep(2)  # brief pause to allow login processing to start
    try:
        WebDriverWait(driver, 20).until(lambda d: BASE_URL in d.current_url)
        print("[3/7] ✅ Login appears successful")
    except TimeoutException:
        print("⚠️ Warning: login may not have completed successfully. Check selectors or credentials.")


def click_area_results_test_menu_span(
    driver,
    target_text: Optional[str] = None,
    timeout: int = 20,
) -> dict:
    """
    Test helper for the new area results flow.

    It clicks a visible ABP/LeptonX menu span after login. If target_text is
    provided it must match the span's normalized text; otherwise the first
    visible span matching the shared menu classes is clicked.
    """
    print("[4/7] Waiting for area results test menu span")
    selector = "span.lpx-menu-item-text.hidden-in-hover-trigger"
    target_text_normalized = " ".join(target_text.split()) if target_text else None

    def find_clickable_span(driver):
        return driver.execute_script(
            r"""
            const selector = arguments[0];
            const targetText = arguments[1];
            const spans = Array.from(document.querySelectorAll(selector));
            return spans.find((span) => {
                const rect = span.getBoundingClientRect();
                const text = (span.textContent || "").replace(/\s+/g, " ").trim();
                if (rect.width <= 0 || rect.height <= 0) return false;
                if (targetText && text !== targetText) return false;
                return true;
            }) || null;
            """,
            selector,
            target_text_normalized,
        )

    span = WebDriverWait(driver, timeout).until(find_clickable_span)
    clicked_text = driver.execute_script(
        "return (arguments[0].textContent || '').replace(/\\s+/g, ' ').trim();",
        span,
    )
    old_url = driver.current_url
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", span)
    try:
        span.click()
    except (ElementClickInterceptedException, StaleElementReferenceException, WebDriverException):
        span = WebDriverWait(driver, 5).until(find_clickable_span)
        driver.execute_script("arguments[0].click();", span)

    print(f"[4/7] Clicked area results test menu span: {clicked_text or '(blank)'}")
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        WebDriverWait(driver, timeout).until(
            lambda d: d.current_url != old_url
            or not d.execute_script(
                """
                const loading = document.querySelector(
                    ".lpx-loader, .ngx-spinner-overlay, .k-loading-mask, .spinner-border"
                );
                return Boolean(loading && loading.offsetParent !== null);
                """
            )
        )
    except TimeoutException:
        print("[4/7] Timed out waiting for visible navigation change after menu click")

    return {
        "clicked_text": clicked_text,
        "start_url": old_url,
        "current_url": driver.current_url,
        "title": driver.title,
    }


def format_area_results_test_date(value: Optional[date]) -> str:
    delivery_date = value or datetime.now().date()
    return delivery_date.strftime("%Y/%m/%d")


def find_area_results_field_input(driver, label: str):
    return driver.execute_script(
        r"""
        const labelText = arguments[0].toLowerCase();
        const normalize = (value) => (value || "")
            .replace(/\*/g, "")
            .replace(/\s+/g, " ")
            .trim()
            .toLowerCase();
        const labels = Array.from(document.querySelectorAll("label, .form-label, div, span"))
            .filter((element) => normalize(element.textContent) === labelText);

        for (const label of labels) {
            let node = label;
            for (let depth = 0; node && depth < 6; depth += 1) {
                const input = node.querySelector(
                    "input.k-input-inner:not([type='hidden']), input:not([type='hidden'])"
                );
                if (input) return input;
                node = node.parentElement;
            }
        }
        return null;
        """,
        label,
    )


def set_area_results_input_by_label(driver, label: str, value: str, timeout: int = 20):
    print(f"[5/7] Setting {label}: {value}")

    def field_input_exists(driver):
        return find_area_results_field_input(driver, label)

    input_element = WebDriverWait(driver, timeout).until(field_input_exists)
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
    input_element.click()
    input_element.send_keys(Keys.CONTROL, "a")
    input_element.send_keys(value)
    input_element.send_keys(Keys.TAB)

    driver.execute_script(
        """
        const input = arguments[0];
        const value = arguments[1];
        const setter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype,
            "value"
        ).set;
        setter.call(input, value);
        input.dispatchEvent(new Event("input", { bubbles: true }));
        input.dispatchEvent(new Event("change", { bubbles: true }));
        input.dispatchEvent(new Event("blur", { bubbles: true }));
        """,
        input_element,
        value,
    )
    return input_element


def select_area_results_category(driver, value: str, timeout: int = 20):
    print(f"[5/7] Selecting Category: {value}")
    input_element = WebDriverWait(driver, timeout).until(
        lambda d: find_area_results_field_input(d, "Category")
    )
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", input_element)
    input_element.click()
    input_element.send_keys(Keys.CONTROL, "a")
    input_element.send_keys(value)

    def visible_option(driver):
        return driver.execute_script(
            r"""
            const target = arguments[0].toLowerCase();
            const options = Array.from(document.querySelectorAll(
                ".k-list-item, .k-item, [role='option']"
            ));
            return options.find((option) => {
                const rect = option.getBoundingClientRect();
                const text = (option.textContent || "").replace(/\s+/g, " ").trim().toLowerCase();
                return rect.width > 0 && rect.height > 0 && text === target;
            }) || null;
            """,
            value,
        )

    try:
        option = WebDriverWait(driver, 5).until(visible_option)
        driver.execute_script("arguments[0].click();", option)
    except TimeoutException:
        input_element.send_keys(Keys.ENTER)
        input_element.send_keys(Keys.TAB)

    driver.execute_script(
        """
        const input = arguments[0];
        const value = arguments[1];
        const setter = Object.getOwnPropertyDescriptor(
            window.HTMLInputElement.prototype,
            "value"
        ).set;
        setter.call(input, value);
        input.dispatchEvent(new Event("input", { bubbles: true }));
        input.dispatchEvent(new Event("change", { bubbles: true }));
        input.dispatchEvent(new Event("blur", { bubbles: true }));
        """,
        input_element,
        value,
    )
    return input_element


def click_area_results_constrained_toggle(driver, timeout: int = 20) -> dict:
    print("[5/7] Finding constrained area result toggle")

    def find_toggle(driver):
        return driver.execute_script(
            r"""
            const normalize = (value) => (value || "")
                .replace(/\*/g, "")
                .replace(/\s+/g, " ")
                .trim()
                .toLowerCase();
            const labels = Array.from(document.querySelectorAll("label, .form-label, div, span"))
                .filter((element) => normalize(element.textContent) === "constrained per area result");

            for (const label of labels) {
                let node = label;
                for (let depth = 0; node && depth < 6; depth += 1) {
                    const toggle = node.querySelector(
                        ".k-switch, .k-switch-track, [role='switch'], input[type='checkbox']"
                    );
                    if (toggle) return toggle;
                    node = node.parentElement;
                }
            }
            return document.querySelector(
                ".k-switch, .k-switch-track, [role='switch'], input[type='checkbox']"
            );
            """
        )

    toggle = WebDriverWait(driver, timeout).until(find_toggle)
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", toggle)
    driver.execute_script("arguments[0].click();", toggle)
    time.sleep(3)
    driver.execute_script("arguments[0].click();", toggle)
    time.sleep(3)
    return {"constrained_toggle_sequence": ["on", "off"]}


def find_area_results_constrained_toggle(driver):
    return driver.execute_script(
        r"""
        const normalize = (value) => (value || "")
            .replace(/\*/g, "")
            .replace(/\s+/g, " ")
            .trim()
            .toLowerCase();
        const labels = Array.from(document.querySelectorAll("label, .form-label, div, span"))
            .filter((element) => normalize(element.textContent) === "constrained per area result");

        for (const label of labels) {
            let node = label;
            for (let depth = 0; node && depth < 6; depth += 1) {
                const toggle = node.querySelector(
                    ".k-switch, [role='switch'], input[type='checkbox'], .k-switch-track"
                );
                if (toggle) return toggle.closest(".k-switch") || toggle;
                node = node.parentElement;
            }
        }
        const fallback = document.querySelector(
            ".k-switch, [role='switch'], input[type='checkbox'], .k-switch-track"
        );
        return fallback ? (fallback.closest(".k-switch") || fallback) : null;
        """
    )


def get_area_results_constrained_toggle_state(driver, toggle) -> Optional[bool]:
    return driver.execute_script(
        """
        const toggle = arguments[0];
        const input = toggle.matches("input[type='checkbox']")
            ? toggle
            : toggle.querySelector("input[type='checkbox']");
        if (input) return Boolean(input.checked);

        const ariaChecked = toggle.getAttribute("aria-checked")
            || (toggle.querySelector("[aria-checked]") || {}).getAttribute?.("aria-checked");
        if (ariaChecked === "true") return true;
        if (ariaChecked === "false") return false;

        const className = toggle.getAttribute("class") || "";
        if (/k-switch-on|k-selected|k-checked/.test(className)) return true;
        if (/k-switch-off/.test(className)) return false;
        return null;
        """,
        toggle,
    )


def set_area_results_constrained_toggle(
    driver,
    enabled: bool,
    timeout: int = 20,
) -> dict:
    label = "on" if enabled else "off"
    print(f"[5/7] Setting constrained area result toggle {label}")
    toggle = WebDriverWait(driver, timeout).until(find_area_results_constrained_toggle)
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", toggle)
    before_state = get_area_results_constrained_toggle_state(driver, toggle)
    should_click = before_state is not enabled
    if before_state is None and enabled is False:
        should_click = False
    if should_click:
        driver.execute_script("arguments[0].click();", toggle)
        time.sleep(2)
    after_state = get_area_results_constrained_toggle_state(driver, toggle)
    print(
        "[5/7] Constrained toggle state: "
        f"before={before_state}, requested={enabled}, after={after_state}"
    )
    return {
        "requested_constrained": enabled,
        "constrained_toggle_before": before_state,
        "constrained_toggle_after": after_state,
    }


def click_area_results_search_button(driver, timeout: int = 20) -> dict:
    print("[6/7] Clicking area results Search button")
    old_url = driver.current_url
    old_signature = driver.execute_script(
        "return document.body ? document.body.innerText.slice(0, 2000) : '';"
    )

    def find_search_button(driver):
        return driver.execute_script(
            r"""
            const buttons = Array.from(document.querySelectorAll(
                "button[title='Search'], button[data-cy='Search'], button"
            ));
            return buttons.find((button) => {
                const rect = button.getBoundingClientRect();
                const text = (button.textContent || "").replace(/\s+/g, " ").trim().toLowerCase();
                const title = (button.getAttribute("title") || "").trim().toLowerCase();
                const dataCy = (button.getAttribute("data-cy") || "").trim().toLowerCase();
                return rect.width > 0
                    && rect.height > 0
                    && !button.disabled
                    && (
                        title === "search"
                        || dataCy === "search"
                        || text === "search"
                    );
            }) || null;
            """
        )

    button = WebDriverWait(driver, timeout).until(find_search_button)
    button_info = driver.execute_script(
        r"""
        const button = arguments[0];
        const rect = button.getBoundingClientRect();
        return {
            text: (button.textContent || "").replace(/\s+/g, " ").trim(),
            title: button.getAttribute("title"),
            dataCy: button.getAttribute("data-cy"),
            disabled: Boolean(button.disabled),
            className: button.getAttribute("class"),
            x: rect.x,
            y: rect.y,
            width: rect.width,
            height: rect.height
        };
        """,
        button,
    )
    print(f"[6/7] Found Search button: {button_info}")
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
    try:
        button.click()
        click_method = "selenium_click"
    except (ElementClickInterceptedException, StaleElementReferenceException, WebDriverException):
        button = WebDriverWait(driver, 5).until(find_search_button)
        driver.execute_script(
            """
            const button = arguments[0];
            button.dispatchEvent(new MouseEvent('mousedown', { bubbles: true, cancelable: true, view: window }));
            button.dispatchEvent(new MouseEvent('mouseup', { bubbles: true, cancelable: true, view: window }));
            button.dispatchEvent(new MouseEvent('click', { bubbles: true, cancelable: true, view: window }));
            """,
            button,
        )
        click_method = "javascript_mouse_events"
    print(f"[6/7] Search button click dispatched via {click_method}")

    redirected = False
    content_changed = False
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: (
                d.current_url != old_url
                or d.execute_script(
                    "return document.body ? document.body.innerText.slice(0, 2000) : '';"
                )
                != old_signature
                or d.execute_script(
                    """
                    const loading = document.querySelector(
                        ".k-loading-mask, .k-i-loading, .k-loading-image, .spinner-border, .ngx-spinner-overlay"
                    );
                    return Boolean(loading && loading.offsetParent !== null);
                    """
                )
            )
        )
    except TimeoutException:
        print("[6/7] No URL/content/loading change detected after Search click")

    time.sleep(5)
    new_url = driver.current_url
    new_signature = driver.execute_script(
        "return document.body ? document.body.innerText.slice(0, 2000) : '';"
    )
    redirected = new_url != old_url
    content_changed = new_signature != old_signature
    print(
        "[6/7] Search result state: "
        f"redirected={redirected}, content_changed={content_changed}, "
        f"current_url={new_url}"
    )
    return {
        "search_clicked": True,
        "search_click_method": click_method,
        "search_button": button_info,
        "search_start_url": old_url,
        "search_current_url": new_url,
        "search_redirected": redirected,
        "search_content_changed": content_changed,
    }


def click_area_results_select_schedule_button(driver, timeout: int = 20) -> dict:
    print("[5/7] Clicking Select a Schedule button")

    def find_select_schedule_button(driver):
        return driver.execute_script(
            r"""
            const buttons = Array.from(document.querySelectorAll(
                "button[data-cy='Select a Schedule'], button[title='Select a Schedule'], button"
            ));
            return buttons.find((button) => {
                const rect = button.getBoundingClientRect();
                const text = (button.textContent || "").replace(/\s+/g, " ").trim().toLowerCase();
                const title = (button.getAttribute("title") || "").trim().toLowerCase();
                const dataCy = (button.getAttribute("data-cy") || "").trim().toLowerCase();
                return rect.width > 0
                    && rect.height > 0
                    && !button.disabled
                    && (
                        title === "select a schedule"
                        || dataCy === "select a schedule"
                        || text === "select a schedule"
                    );
            }) || null;
            """
        )

    button = WebDriverWait(driver, timeout).until(find_select_schedule_button)
    button_info = driver.execute_script(
        r"""
        const button = arguments[0];
        const rect = button.getBoundingClientRect();
        return {
            text: (button.textContent || "").replace(/\s+/g, " ").trim(),
            title: button.getAttribute("title"),
            dataCy: button.getAttribute("data-cy"),
            disabled: Boolean(button.disabled),
            className: button.getAttribute("class"),
            x: rect.x,
            y: rect.y,
            width: rect.width,
            height: rect.height
        };
        """,
        button,
    )
    print(f"[5/7] Found Select a Schedule button: {button_info}")
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
    try:
        button.click()
        click_method = "selenium_click"
    except (ElementClickInterceptedException, StaleElementReferenceException, WebDriverException):
        button = WebDriverWait(driver, 5).until(find_select_schedule_button)
        driver.execute_script("arguments[0].click();", button)
        click_method = "javascript_click"

    WebDriverWait(driver, timeout).until(
        lambda d: find_area_results_field_input(d, "Delivery Day")
    )
    time.sleep(2)
    return {
        "select_schedule_clicked": True,
        "select_schedule_click_method": click_method,
        "select_schedule_button": button_info,
    }


def extract_area_results_handsontable(driver, dataset: str, timeout: int = 20) -> dict:
    print(f"[7/7] Extracting {dataset} area results Handsontable")

    def table_exists(driver):
        return driver.execute_script(
            """
            const table = document.querySelector(".ht_master table.htCore");
            if (!table) return null;
            const headers = Array.from(table.querySelectorAll("thead th"))
                .map((cell) => (cell.textContent || "").trim())
                .filter(Boolean);
            const rows = Array.from(table.querySelectorAll("tbody tr"));
            return headers.length >= 2 && rows.length > 0 ? table : null;
            """
        )

    WebDriverWait(driver, timeout).until(table_exists)
    table_data = driver.execute_script(
        r"""
        const dataset = arguments[0];
        const table = document.querySelector(".ht_master table.htCore");
        const normalize = (value) => (value || "").replace(/\s+/g, " ").trim();
        const parseNumber = (value) => {
            const text = normalize(value).replace(/,/g, "");
            if (!text) return null;
            const parsed = Number(text);
            return Number.isFinite(parsed) ? parsed : null;
        };
        const normalizeDate = (value) => normalize(value).replace(/\//g, "-");
        const hourFromProduct = (product) => {
            const match = normalize(product).match(/^(\d{2})-(\d{2})$/);
            if (!match) return null;
            return Number(match[1]) + 1;
        };
        const headers = Array.from(table.querySelectorAll("thead th"))
            .map((cell) => normalize(cell.textContent));
        const dateHeaders = headers.slice(1).map((label) => ({
            label,
            date: normalizeDate(label)
        }));
        const rows = Array.from(table.querySelectorAll("tbody tr")).map((row) => {
            const cells = Array.from(row.querySelectorAll("td, th"))
                .map((cell) => normalize(cell.textContent));
            const product = cells[0] || "";
            const values = {};
            dateHeaders.forEach((header, index) => {
                values[header.date] = parseNumber(cells[index + 1]);
            });
            return { product, values };
        }).filter((row) => row.product);

        const summaryNames = new Set(["min", "max", "avg", "nett"]);
        const hourlyRecords = [];
        const summaryRecords = [];
        rows.forEach((row) => {
            const normalizedProduct = row.product.toLowerCase();
            for (const [date, value] of Object.entries(row.values)) {
                if (summaryNames.has(normalizedProduct)) {
                    summaryRecords.push({
                        dataset,
                        date,
                        metric: row.product,
                        value
                    });
                } else {
                    hourlyRecords.push({
                        dataset,
                        date,
                        product: row.product,
                        hour: hourFromProduct(row.product),
                        value
                    });
                }
            }
        });

        return {
            dataset,
            columns: dateHeaders,
            rows,
            hourly_records: hourlyRecords,
            summary_records: summaryRecords,
            row_count: rows.length,
            hourly_record_count: hourlyRecords.length,
            summary_record_count: summaryRecords.length
        };
        """,
        dataset,
    )
    print(
        f"[7/7] Extracted {dataset} table: "
        f"{table_data['row_count']} rows, "
        f"{table_data['hourly_record_count']} hourly cells, "
        f"{table_data['summary_record_count']} summary cells"
    )
    return table_data


def search_and_extract_area_results_dataset(
    driver,
    dataset: str,
    constrained: bool,
    timeout: int = 20,
) -> dict:
    toggle_result = set_area_results_constrained_toggle(
        driver,
        constrained,
        timeout=timeout,
    )
    search_result = click_area_results_search_button(driver, timeout=timeout)
    table_result = extract_area_results_handsontable(driver, dataset, timeout=timeout)
    return {
        "dataset": dataset,
        "constrained": constrained,
        **toggle_result,
        **search_result,
        "table": table_result,
    }


def run_area_results_test(
    target_text: Optional[str] = None,
    timeout: int = 20,
    delivery_date: Optional[date] = None,
) -> dict:
    username, password = load_config()
    with tempfile.TemporaryDirectory(prefix="sapp-area-results-test-") as temp_download_dir:
        download_dir = Path(temp_download_dir)
        driver = create_driver(download_dir, headless=False)
        try:
            login(driver, username, password)
            print(f"[4/7] Navigating directly to area results test URL: {AREA_RESULTS_TEST_URL}")
            driver.get(AREA_RESULTS_TEST_URL)
            WebDriverWait(driver, timeout).until(
                lambda d: d.execute_script("return document.readyState") == "complete"
            )
            time.sleep(5)
            delivery_day = format_area_results_test_date(delivery_date)
            set_area_results_input_by_label(driver, "Delivery Day", delivery_day, timeout=timeout)
            select_area_results_category(driver, "Price in USD", timeout=timeout)
            unconstrained_result = search_and_extract_area_results_dataset(
                driver,
                "unconstrained",
                constrained=False,
                timeout=timeout,
            )
            select_schedule_result = click_area_results_select_schedule_button(
                driver,
                timeout=timeout,
            )
            set_area_results_input_by_label(driver, "Delivery Day", delivery_day, timeout=timeout)
            select_area_results_category(driver, "Price in USD", timeout=timeout)
            constrained_result = search_and_extract_area_results_dataset(
                driver,
                "constrained",
                constrained=True,
                timeout=timeout,
            )
            return {
                "status": "success",
                "target_text": target_text,
                "target_url": AREA_RESULTS_TEST_URL,
                "delivery_day": delivery_day,
                "category": "Price in USD",
                "returned_dates": constrained_result["table"]["columns"],
                **select_schedule_result,
                "datasets": {
                    "unconstrained": unconstrained_result,
                    "constrained": constrained_result,
                },
                "current_url": driver.current_url,
                "title": driver.title,
            }
        finally:
            time.sleep(30)
            driver.quit()


def navigate_to_inbox(driver):
    print("[4/7] 📥 Navigating to inbox")
    driver.get(INBOX_URL)
    wait_for_inbox_grid_to_settle(driver, timeout=INBOX_GRID_READY_TIMEOUT)
    print("[4/7] ✅ Inbox page loaded")


def xpath_literal(value: str) -> str:
    if "'" not in value:
        return f"'{value}'"
    if '"' not in value:
        return f'"{value}"'
    parts = value.split("'")
    return "concat(" + ', "\'", '.join(f"'{part}'" for part in parts) + ")"


def wait_for_inbox_grid_to_settle(driver, timeout: int = INBOX_GRID_READY_TIMEOUT):
    try:
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located(
                (By.CSS_SELECTOR, INBOX_LOADING_SELECTOR)
            )
        )
        WebDriverWait(driver, timeout).until(
            lambda d: d.execute_script(
                """
                const grid = document.querySelector(arguments[0]);
                if (!grid) return false;
                const loading = document.querySelector(arguments[1]);
                if (loading && loading.offsetParent !== null) return false;
                return Boolean(
                    grid.querySelector("tbody tr") ||
                    /no records|no data|empty/i.test(grid.innerText)
                );
                """,
                INBOX_GRID_SELECTOR,
                INBOX_LOADING_SELECTOR,
            )
        )
    except TimeoutException:
        pass


def inbox_grid_signature(driver) -> str:
    try:
        return driver.execute_script(
            """
            const selectedPage = document.querySelector(
                ".k-pager-numbers .k-selected, .k-pager-numbers .k-state-selected, [aria-current='page']"
            )?.textContent?.trim() || "";
            const grid = document.querySelector(arguments[0]) || document.body;
            return selectedPage + "|" + grid.innerText.slice(0, 2000);
            """,
            INBOX_GRID_SELECTOR,
        )
    except WebDriverException:
        return ""


def get_current_inbox_page_number(driver) -> Optional[int]:
    try:
        value = driver.execute_script(
            """
            const input = document.querySelector(arguments[0]);
            const selected = document.querySelector(
                ".k-pager-numbers .k-selected, .k-pager-numbers .k-state-selected, [aria-current='page']"
            );
            return (
                input?.getAttribute("aria-valuenow") ||
                input?.value ||
                selected?.textContent?.trim() ||
                null
            );
            """,
            PAGE_NUMBER_INPUT_SELECTOR,
        )
    except WebDriverException:
        return None

    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return None


def find_visible_subject_titles(driver) -> list[str]:
    try:
        return driver.execute_script(
            """
            return Array.from(document.querySelectorAll("span[title]"))
                .filter((element) => {
                    const rect = element.getBoundingClientRect();
                    return rect.width > 0 && rect.height > 0;
                })
                .map((element) => (element.getAttribute("title") || "").trim())
                .filter(Boolean);
            """
        )
    except WebDriverException:
        return []


def find_matching_subject_title_on_current_page(
    driver,
    job: SappExtractionJob,
    delivery_date: date,
) -> Optional[str]:
    for subject in find_visible_subject_titles(driver):
        if subject_matches_job_and_date(subject, job, delivery_date):
            return subject
    return None


def find_matching_subjects_on_current_page(
    driver,
    job: SappExtractionJob,
    pending_dates: list[date],
) -> dict[str, date]:
    matches = {}
    visible_subjects = find_visible_subject_titles(driver)
    for delivery_date in pending_dates:
        for subject in visible_subjects:
            if subject in matches:
                continue
            if subject_matches_job_and_date(subject, job, delivery_date):
                matches[subject] = delivery_date
                break
    return matches


def find_matching_job_subjects_on_current_page(
    driver,
    jobs: list[SappExtractionJob],
    pending_items: list[tuple[SappExtractionJob, date]],
) -> dict[str, tuple[SappExtractionJob, date]]:
    matches = {}
    visible_subjects = find_visible_subject_titles(driver)
    for job, delivery_date in pending_items:
        for subject in visible_subjects:
            if subject in matches:
                continue
            if subject_matches_job_and_date(subject, job, delivery_date):
                matches[subject] = (job, delivery_date)
                break
    return matches


def extract_job_delivery_dates_from_current_page(
    driver,
    job: SappExtractionJob,
) -> list[date]:
    prefix = job.subject_template.split("{delivery_date}")[0]
    extracted_dates = []
    for subject in find_visible_subject_titles(driver):
        if not subject.startswith(prefix):
            continue
        candidate = subject[len(prefix) :].strip()
        parsed_date = parse_delivery_date_value(candidate)
        if parsed_date is not None:
            extracted_dates.append(parsed_date)
    return extracted_dates


def click_message(driver, target_subject: str):
    subject_xpath = f"//span[@title={xpath_literal(target_subject)}]"
    try:
        message_span = WebDriverWait(driver, MESSAGE_CLICK_TIMEOUT).until(
            EC.element_to_be_clickable((By.XPATH, subject_xpath))
        )
        try:
            message_span.click()
        except WebDriverException:
            driver.execute_script("arguments[0].click();", message_span)
    except TimeoutException:
        message_span = find_message_on_current_page(driver, target_subject)
        if message_span is None:
            raise
        driver.execute_script("arguments[0].click();", message_span)


def find_enabled_next_page_button(driver):
    for button in driver.find_elements(By.CSS_SELECTOR, NEXT_PAGE_SELECTOR):
        try:
            class_name = button.get_attribute("class") or ""
            if (
                button.is_displayed()
                and button.is_enabled()
                and button.get_attribute("aria-disabled") != "true"
                and button.get_attribute("disabled") is None
                and "k-disabled" not in class_name
            ):
                return button
        except StaleElementReferenceException:
            continue
    return None


def click_next_inbox_page(driver, page_number: int, timeout: int = INBOX_PAGE_CHANGE_TIMEOUT):
    old_signature = inbox_grid_signature(driver)
    print(f"[5/7] Attempting next-page navigation from page {page_number} with timeout={timeout}s")

    def click_next(driver):
        button = find_enabled_next_page_button(driver)
        if button is None:
            return False
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", button)
        try:
            button.click()
        except (ElementClickInterceptedException, StaleElementReferenceException, WebDriverException):
            button = find_enabled_next_page_button(driver)
            if button is None:
                return False
            driver.execute_script("arguments[0].click();", button)
        return True

    if not click_next(driver):
        raise RuntimeError(f"Could not click the inbox next-page button from page {page_number}.")

    try:
        WebDriverWait(driver, timeout).until(
            lambda d: inbox_grid_signature(d) != old_signature
        )
    except TimeoutException as exc:
        raise RuntimeError(
            f"Timed out waiting for inbox page {page_number + 1} to load after clicking next."
        ) from exc

    wait_for_inbox_grid_to_settle(driver)


def go_to_inbox_page(driver, page_start: int):
    if page_start <= 1:
        return

    print(f"[5/7] Jumping to inbox page {page_start}")
    wait_for_inbox_grid_to_settle(driver, timeout=INBOX_GRID_READY_TIMEOUT)
    old_signature = inbox_grid_signature(driver)

    try:
        print(f"[5/7] Trying direct page jump to {page_start} with timeout={INBOX_PAGE_CHANGE_TIMEOUT}s")
        page_input = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, PAGE_NUMBER_INPUT_SELECTOR))
        )
        driver.execute_script(
            "arguments[0].scrollIntoView({block: 'center'});",
            page_input,
        )
        page_input.click()
        page_input.send_keys(Keys.CONTROL, "a")
        page_input.send_keys(str(page_start))
        page_input.send_keys(Keys.ENTER)

        WebDriverWait(driver, INBOX_PAGE_CHANGE_TIMEOUT).until(
            lambda d: (
                get_current_inbox_page_number(d) == page_start
                or inbox_grid_signature(d) != old_signature
            )
        )
        wait_for_inbox_grid_to_settle(driver)
        current_page = get_current_inbox_page_number(driver)
        if current_page == page_start or inbox_grid_signature(driver) != old_signature:
            print(f"[5/7] Inbox page {page_start} loaded")
            return
    except (TimeoutException, WebDriverException):
        print(
            f"[5/7] Could not jump directly to inbox page {page_start}; "
            "falling back to next-page navigation"
        )

    current_page = get_current_inbox_page_number(driver) or 1
    while current_page < page_start:
        click_next_inbox_page(driver, current_page)
        current_page += 1


def find_message_and_open(
    driver,
    job: SappExtractionJob,
    delivery_date: date,
    page_start: int = 1,
):
    target_subject = job.build_subject(delivery_date)
    print(f"[5/7] Looking for message with subject: {target_subject}")
    go_to_inbox_page(driver, page_start)

    for page_number in range(page_start, page_start + MAX_INBOX_PAGES_TO_SEARCH):
        wait_for_inbox_grid_to_settle(driver, timeout=INBOX_GRID_READY_TIMEOUT)
        matched_subject = find_matching_subject_title_on_current_page(
            driver,
            job,
            delivery_date,
        )
        if matched_subject is not None:
            click_message(driver, matched_subject)
            print(f"[5/7] Opened target message on inbox page {page_number}")
            return

        next_page_button = find_enabled_next_page_button(driver)
        if next_page_button is None:
            break

        print(f"[5/7] Message not on page {page_number}; moving to next inbox page")
        click_next_inbox_page(driver, page_number)

    raise RuntimeError(
        f"Target message not found after searching up to {MAX_INBOX_PAGES_TO_SEARCH} "
        f"inbox pages from page {page_start}: {target_subject}"
    )

def wait_for_new_download_file(
    download_dir: Path,
    existing_names: set,
    extension: str = ".xlsx",
    timeout: int = 30,
) -> Path:
    print("[6/7] ⏳ Waiting for the downloaded Excel file to appear")
    deadline = time.time() + timeout
    while time.time() < deadline:
        files = [p for p in download_dir.glob(f"*{extension}") if p.name not in existing_names]
        if files:
            latest = max(files, key=lambda p: p.stat().st_mtime)
            if latest.stat().st_size > 0:
                print(f"[6/7] 📥 Download detected: {latest.name}")
                return latest
        time.sleep(1)
    raise RuntimeError("Timed out waiting for the downloaded Excel file.")


def find_latest_download_file(download_dir: Path) -> Path:
    files = list(download_dir.glob("*.xlsx"))
    if not files:
        raise RuntimeError("No Excel files found in the download folder.")
    return max(files, key=lambda p: p.stat().st_mtime)


def extract_data_from_file(file_path: Path):
    print(f"[7/7] 📄 Extracting data from downloaded file: {file_path.name}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    delivery_date = None
    for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
        if not row:
            continue
        for idx, value in enumerate(row):
            if isinstance(value, str) and value.strip() == "Delivery Date:":
                candidate = row[idx + 1] if idx + 1 < len(row) else None
                if isinstance(candidate, datetime):
                    delivery_date = candidate.date()
                elif isinstance(candidate, str):
                    try:
                        delivery_date = datetime.fromisoformat(candidate).date()
                    except ValueError:
                        continue
                break
        if delivery_date is not None:
            break

    if delivery_date is None:
        raise RuntimeError("Could not find the delivery date in the downloaded file.")

    header_row_index = None
    hour_col = purchase_col = sales_col = price_col = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        if (
            "Hour" in row
            and "Area Purchase (MW)" in row
            and "Area Sales (MW)" in row
            and "Area Price (USD/MWh)" in row
        ):
            header_row_index = row_idx
            hour_col = row.index("Hour")
            purchase_col = row.index("Area Purchase (MW)")
            sales_col = row.index("Area Sales (MW)")
            price_col = row.index("Area Price (USD/MWh)")
            break

    if header_row_index is None:
        raise RuntimeError("Could not find the hourly data header row in the downloaded file.")

    extracted_rows = []
    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not row or row[hour_col] is None:
            continue
        hour_text = str(row[hour_col]).strip()
        hour_value = hour_text.split("-")[0] if "-" in hour_text else hour_text
        area_purchase = row[purchase_col] if purchase_col < len(row) else None
        area_sales = row[sales_col] if sales_col < len(row) else None
        area_price = row[price_col] if price_col < len(row) else None
        extracted_rows.append([delivery_date.isoformat(), hour_value, area_purchase, area_sales, area_price])

    if not extracted_rows:
        raise RuntimeError("No hourly rows were extracted from the downloaded file.")

    output_wb = openpyxl.Workbook()
    output_ws = output_wb.active
    output_ws.title = "Extracted Data"
    output_ws.append(["Date", "Hour", "Area Purchase (MW)", "Area Sales (MW)", "Area Price (USD/MWh)"])
    for row in extracted_rows:
        output_ws.append(row)

    output_path = file_path.with_name(f"extracted_{file_path.stem}.xlsx")
    output_wb.save(output_path)
    print(f"💾 Saved extracted data to: {output_path}")
    return output_path


def extract_constrained_area_results(
    file_path: Path,
    data_source: str = CONSTRAINED_AREA_DATA_SOURCE,
) -> list[dict]:
    print(f"[7/7] Extracting SAPP rows from downloaded file: {file_path.name}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    delivery_date = None
    for row in sheet.iter_rows(min_row=1, max_row=50, values_only=True):
        if not row:
            continue
        for idx, value in enumerate(row):
            if isinstance(value, str) and value.strip() == "Delivery Date:":
                candidate = row[idx + 1] if idx + 1 < len(row) else None
                delivery_date = parse_delivery_date_value(candidate)
                break
        if delivery_date is not None:
            break

    if delivery_date is None:
        raise RuntimeError("Could not find the delivery date in the downloaded file.")

    header_row_index = None
    hour_col = purchase_col = sales_col = price_col = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        if (
            "Hour" in row
            and "Area Purchase (MW)" in row
            and "Area Sales (MW)" in row
            and "Area Price (USD/MWh)" in row
        ):
            header_row_index = row_idx
            hour_col = row.index("Hour")
            purchase_col = row.index("Area Purchase (MW)")
            sales_col = row.index("Area Sales (MW)")
            price_col = row.index("Area Price (USD/MWh)")
            break

    if header_row_index is None:
        raise RuntimeError("Could not find the hourly data header row in the downloaded file.")

    records = []
    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        if not row or hour_col >= len(row) or row[hour_col] is None:
            continue

        hour = parse_hour(row[hour_col])
        if hour is None:
            continue

        area_purchase = row[purchase_col] if purchase_col < len(row) else None
        area_sales = row[sales_col] if sales_col < len(row) else None
        area_price = row[price_col] if price_col < len(row) else None
        records.append(
            {
                "timestamp": delivery_timestamp(delivery_date, hour),
                "delivery_date": delivery_date.isoformat(),
                "hour": hour,
                "hour_label": str(row[hour_col]).strip(),
                "area_purchase_mw": to_float(area_purchase),
                "area_sales_mw": to_float(area_sales),
                "area_price_usd_per_mwh": to_float(area_price),
                "metadata": {
                    "data_source": data_source,
                    "source_file": file_path.name,
                },
                "source_file": file_path.name,
            }
        )

    if not records:
        raise RuntimeError("No hourly rows were extracted from the downloaded file.")

    extracted_hours = {record["hour"] for record in records}
    expected_hours = set(range(1, 25))
    if extracted_hours != expected_hours:
        missing_hours = sorted(expected_hours - extracted_hours)
        extra_hours = sorted(extracted_hours - expected_hours)
        raise RuntimeError(
            "Expected 24 hourly rows for the delivery day, "
            f"but extracted {len(records)}. "
            f"Missing hours: {missing_hours or 'none'}. "
            f"Unexpected hours: {extra_hours or 'none'}."
        )

    return records


def extract_constrained_area_job(file_path: Path, job: SappExtractionJob) -> list[dict]:
    return extract_constrained_area_results(file_path, data_source=job.data_source)


def extract_unconstrained_area_results(
    file_path: Path,
    data_source: str = UNCONSTRAINED_AREA_DATA_SOURCE,
) -> list[dict]:
    print(f"[7/7] Extracting SAPP unconstrained rows from downloaded file: {file_path.name}")
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    sheet = workbook.active

    delivery_date = find_delivery_date_in_sheet(sheet)
    if delivery_date is None:
        raise RuntimeError("Could not find the delivery date in the downloaded file.")

    required_headers = {
        "hour": "hour",
        "total_purchase_volume_mw": "total purchase volume (mw)",
        "total_sales_volume_mw": "total sales volume (mw)",
        "price_usd_per_mwh": "price (usd/mwh)",
        "price_zar_per_mwh": "price (zar/mwh)",
    }
    header_row_index = None
    column_indexes = None
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        normalized = [normalize_excel_header(value) for value in row]
        indexes = {}
        for field_name, header in required_headers.items():
            try:
                indexes[field_name] = normalized.index(header)
            except ValueError:
                indexes = {}
                break
        if indexes:
            header_row_index = row_idx
            column_indexes = indexes
            break

    if header_row_index is None or column_indexes is None:
        raise RuntimeError("Could not find the unconstrained hourly data header row.")

    records = []
    for row in sheet.iter_rows(min_row=header_row_index + 1, values_only=True):
        hour_col = column_indexes["hour"]
        if not row or hour_col >= len(row) or row[hour_col] is None:
            continue

        hour = parse_hour(row[hour_col])
        if hour is None:
            continue

        def cell(field_name: str):
            col_idx = column_indexes[field_name]
            return row[col_idx] if col_idx < len(row) else None

        records.append(
            {
                "timestamp": delivery_timestamp(delivery_date, hour),
                "delivery_date": delivery_date.isoformat(),
                "hour": hour,
                "hour_label": str(row[hour_col]).strip(),
                "total_purchase_volume_mw": to_float(cell("total_purchase_volume_mw")),
                "total_sales_volume_mw": to_float(cell("total_sales_volume_mw")),
                "price_usd_per_mwh": to_float(cell("price_usd_per_mwh")),
                "price_zar_per_mwh": to_float(cell("price_zar_per_mwh")),
                "metadata": {
                    "data_source": data_source,
                    "source_file": file_path.name,
                },
                "source_file": file_path.name,
            }
        )

    if not records:
        raise RuntimeError("No hourly rows were extracted from the downloaded file.")

    extracted_hours = {record["hour"] for record in records}
    expected_hours = set(range(1, 25))
    if extracted_hours != expected_hours:
        missing_hours = sorted(expected_hours - extracted_hours)
        extra_hours = sorted(extracted_hours - expected_hours)
        raise RuntimeError(
            "Expected 24 hourly unconstrained rows for the delivery day, "
            f"but extracted {len(records)}. "
            f"Missing hours: {missing_hours or 'none'}. "
            f"Unexpected hours: {extra_hours or 'none'}."
        )

    return records


def extract_unconstrained_area_job(file_path: Path, job: SappExtractionJob) -> list[dict]:
    return extract_unconstrained_area_results(file_path, data_source=job.data_source)


def extract_participant_portfolio_results(
    file_path: Path,
    data_source: str = PARTICIPANT_PORTFOLIO_DATA_SOURCE,
) -> list[dict]:
    print(
        "[7/7] Extracting SAPP participant portfolio rows from downloaded file: "
        f"{file_path.name}"
    )
    workbook = openpyxl.load_workbook(file_path, data_only=True)

    source_delivery_date = None
    for sheet in workbook.worksheets:
        source_delivery_date = find_portfolio_reference_date_in_sheet(sheet)
        if source_delivery_date is not None:
            break

    if source_delivery_date is None:
        raise RuntimeError("Could not find the delivery date in the downloaded file.")

    dam_required_headers = {
        "hour": "hour",
        "participant_total_area_schedule_mwh": "participant total area schedule (mwh)",
        "area_price_usd_per_mwh": "area price (usd/mwh)",
    }
    fpm_required_headers = {
        "product": "product",
        "participant_total_area_schedule_mw": "participant total area schedule (mw)",
        "area_price_usd_per_mwh": "area price (usd/mwh)",
    }

    records = []

    def build_base_record(
        market: PortfolioMarket,
        delivery_day: date,
        hour: int,
        participant_schedule: Optional[float],
        area_price: Optional[float],
        source_hour_label: str,
        product: Optional[str],
        period_start_date: date,
        period_end_date: date,
        unconstrained_price: Optional[float] = None,
        total_dam_turnover: Optional[float] = None,
    ) -> dict:
        return {
            "timestamp": delivery_timestamp(delivery_day, hour),
            "market": market,
            "delivery_date": delivery_day.isoformat(),
            "source_delivery_date": source_delivery_date.isoformat(),
            "period_start_date": period_start_date.isoformat(),
            "period_end_date": period_end_date.isoformat(),
            "hour": hour,
            "hour_label": hour_label_from_hour(hour),
            "product": product or get_time_of_use_period(delivery_day, hour),
            "participant_total_area_schedule_mwh": participant_schedule,
            "area_price_usd_per_mwh": area_price,
            "unconstrained_market_price_usd_per_mwh": unconstrained_price,
            "total_dam_turnover_mwh": total_dam_turnover,
            "metadata": {
                "data_source": data_source,
                "source_file": file_path.name,
                "market": market,
                "source_hour_label": source_hour_label,
                "source_delivery_date": source_delivery_date.isoformat(),
                "period_start_date": period_start_date.isoformat(),
                "period_end_date": period_end_date.isoformat(),
            },
            "source_file": file_path.name,
        }

    for sheet in workbook.worksheets:
        sheet_rows = list(sheet.iter_rows(values_only=True))
        for row_idx, row in enumerate(sheet_rows, start=1):
            normalized = [normalize_excel_header(value) for value in row]

            dam_indexes = {}
            for field_name, header in dam_required_headers.items():
                try:
                    dam_indexes[field_name] = normalized.index(header)
                except ValueError:
                    dam_indexes = {}
                    break

            if dam_indexes:
                if "unconstrained market price (usd/mwh)" in normalized:
                    dam_indexes["unconstrained_market_price_usd_per_mwh"] = normalized.index(
                        "unconstrained market price (usd/mwh)"
                    )
                if "total dam turnover (mwh)" in normalized:
                    dam_indexes["total_dam_turnover_mwh"] = normalized.index(
                        "total dam turnover (mwh)"
                    )

                market = infer_portfolio_market(sheet, row_idx, "dam")
                period_start_date, period_end_date = period_range_for_portfolio_market(
                    market,
                    source_delivery_date,
                )

                for data_row in sheet_rows[row_idx:]:
                    hour_col = dam_indexes["hour"]
                    if not data_row or hour_col >= len(data_row) or data_row[hour_col] is None:
                        continue

                    raw_hour = str(data_row[hour_col]).strip()
                    if normalize_excel_header(raw_hour) in {"total", "min", "max", "avg"}:
                        break

                    hour = parse_hour(data_row[hour_col])
                    if hour is None:
                        continue

                    def dam_cell(field_name: str):
                        col_idx = dam_indexes.get(field_name)
                        if col_idx is None or col_idx >= len(data_row):
                            return None
                        return data_row[col_idx]

                    delivery_day = source_delivery_date if market == "dam" else period_start_date
                    records.append(
                        build_base_record(
                            market=market,
                            delivery_day=delivery_day,
                            hour=hour,
                            participant_schedule=to_float(
                                dam_cell("participant_total_area_schedule_mwh")
                            ),
                            area_price=to_float(dam_cell("area_price_usd_per_mwh")),
                            source_hour_label=raw_hour,
                            product=get_time_of_use_period(delivery_day, hour),
                            period_start_date=period_start_date,
                            period_end_date=period_end_date,
                            unconstrained_price=to_float(
                                dam_cell("unconstrained_market_price_usd_per_mwh")
                            ),
                            total_dam_turnover=to_float(dam_cell("total_dam_turnover_mwh")),
                        )
                    )
                continue

            fpm_indexes = {}
            for field_name, header in fpm_required_headers.items():
                try:
                    fpm_indexes[field_name] = normalized.index(header)
                except ValueError:
                    fpm_indexes = {}
                    break

            if fpm_indexes:
                market = infer_portfolio_market(sheet, row_idx, "fpm_w")
                if market == "dam":
                    continue
                period_start_date, period_end_date = period_range_for_portfolio_market(
                    market,
                    source_delivery_date,
                )

                product_rows = []
                for data_row in sheet_rows[row_idx:]:
                    product_col = fpm_indexes["product"]
                    if not data_row or product_col >= len(data_row) or data_row[product_col] is None:
                        continue

                    product = normalize_portfolio_product(data_row[product_col])
                    if product is None:
                        first_value = normalize_excel_header(data_row[product_col])
                        if first_value in {"total", "min", "max", "avg"}:
                            break
                        continue

                    def fpm_cell(field_name: str):
                        col_idx = fpm_indexes.get(field_name)
                        if col_idx is None or col_idx >= len(data_row):
                            return None
                        return data_row[col_idx]

                    product_rows.append(
                        {
                            "product": product,
                            "source_hour_label": str(data_row[product_col]).strip(),
                            "participant_total_area_schedule_mwh": to_float(
                                fpm_cell("participant_total_area_schedule_mw")
                            ),
                            "area_price_usd_per_mwh": to_float(
                                fpm_cell("area_price_usd_per_mwh")
                            ),
                        }
                    )

                for current_date in (
                    period_start_date + timedelta(days=offset)
                    for offset in range((period_end_date - period_start_date).days + 1)
                ):
                    for hour in range(1, 25):
                        product = get_time_of_use_period(current_date, hour)
                        matching_row = next(
                            (row_data for row_data in product_rows if row_data["product"] == product),
                            None,
                        )
                        if matching_row is None:
                            continue

                        records.append(
                            build_base_record(
                                market=market,
                                delivery_day=current_date,
                                hour=hour,
                                participant_schedule=matching_row[
                                    "participant_total_area_schedule_mwh"
                                ],
                                area_price=matching_row["area_price_usd_per_mwh"],
                                source_hour_label=matching_row["source_hour_label"],
                                product=product,
                                period_start_date=period_start_date,
                                period_end_date=period_end_date,
                            )
                        )

    if not records:
        raise RuntimeError("No participant portfolio rows were extracted from the downloaded file.")

    deduped_records = {}
    for record in records:
        deduped_records[(record["market"], record["delivery_date"], record["hour"])] = record

    return list(
        sorted(
            deduped_records.values(),
            key=lambda record: (record["market"], record["delivery_date"], record["hour"]),
        )
    )


def extract_participant_portfolio_job(file_path: Path, job: SappExtractionJob) -> list[dict]:
    return extract_participant_portfolio_results(file_path, data_source=job.data_source)


def extract_trading_invoice_results(
    file_path: Path,
    data_source: str = TRADING_INVOICE_DATA_SOURCE,
) -> list[dict]:
    print(
        "[7/7] Extracting SAPP trading invoice rows from downloaded file: "
        f"{file_path.name}"
    )
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    summary_sheet = None
    for sheet in workbook.worksheets:
        labels = {
            str(row[0]).strip()
            for row in sheet.iter_rows(values_only=True)
            if row and row[0] is not None
        }
        if "Net Amount Traded" in labels and "Total Amount Due" in labels:
            summary_sheet = sheet
            break

    if summary_sheet is None:
        raise RuntimeError("Could not find the invoice summary sheet.")

    delivery_date = find_delivery_date_in_sheet(summary_sheet)
    if delivery_date is None:
        raise RuntimeError("Could not find the delivery date in the downloaded file.")

    mwh_col, total_col, price_col = find_summary_value_columns(summary_sheet)
    market_sections = {
        "fpm_m": build_invoice_market_section(
            summary_sheet,
            "FPM-M",
            mwh_col,
            total_col,
            price_col,
        ),
        "fpm_w": build_invoice_market_section(
            summary_sheet,
            "FPM-W",
            mwh_col,
            total_col,
            price_col,
        ),
        "dam": build_invoice_market_section(
            summary_sheet,
            "DAM",
            mwh_col,
            total_col,
            price_col,
        ),
        "idm": build_invoice_market_section(
            summary_sheet,
            "IDM",
            mwh_col,
            total_col,
            price_col,
        ),
    }
    active_market_sections = {
        market: section
        for market, section in market_sections.items()
        if has_invoice_section_activity(section)
    }
    balancing_market = {
        "up_regulation_activated": get_summary_row_values(
            summary_sheet,
            "BM Total Up Regulation Activated",
            mwh_col,
            total_col,
            price_col,
        ),
        "down_regulation_supplied": get_summary_row_values(
            summary_sheet,
            "BM Total Down Regulation Supplied",
            mwh_col,
            total_col,
            price_col,
        ),
        "up_regulation_supplied": get_summary_row_values(
            summary_sheet,
            "BM Total Up Regulation Supplied",
            mwh_col,
            total_col,
            price_col,
        ),
        "down_regulation_activated": get_summary_row_values(
            summary_sheet,
            "BM Total Down Regulation Activated",
            mwh_col,
            total_col,
            price_col,
        ),
    }

    total_purchases = get_summary_row_values(
        summary_sheet,
        "Total Purchases",
        mwh_col,
        total_col,
        price_col,
    )
    total_sales = get_summary_row_values(
        summary_sheet,
        "Total Sales",
        mwh_col,
        total_col,
        price_col,
    )
    net_amount_traded = get_summary_row_values(
        summary_sheet,
        "Net Amount Traded",
        mwh_col,
        total_col,
        price_col,
    )["amount_usd"]
    admin_fee = get_summary_row_values(
        summary_sheet,
        "Total Admin Fee",
        mwh_col,
        total_col,
        price_col,
    )
    wheeling_fee = get_summary_row_values(
        summary_sheet,
        "Total Wheeling Fee",
        mwh_col,
        total_col,
        price_col,
    )["amount_usd"]
    losses_fee = get_summary_row_values(
        summary_sheet,
        "Total Losses Fee",
        mwh_col,
        total_col,
        price_col,
    )["amount_usd"]
    total_fees = get_summary_row_values(
        summary_sheet,
        "Total Fees",
        mwh_col,
        total_col,
        price_col,
    )["amount_usd"]
    total_amount_due = get_summary_row_values(
        summary_sheet,
        "Total Amount Due",
        mwh_col,
        total_col,
        price_col,
    )["amount_usd"]

    gross_total_mwh = sum(
        section["mwh"] or 0.0 for section in active_market_sections.values()
    )
    gross_total_amount = sum(
        section["amount_usd"] or 0.0 for section in active_market_sections.values()
    )
    total_expenditure = (
        gross_total_amount
        + (wheeling_fee or 0.0)
        + (admin_fee["amount_usd"] or 0.0)
    )
    confirmed_trade_type = (
        "PURCHASE"
        if (net_amount_traded or 0.0) > 0
        else "SALE"
        if (net_amount_traded or 0.0) < 0
        else "NONE"
    )

    record = {
        "timestamp": datetime.combine(delivery_date, datetime.min.time()),
        "delivery_date": delivery_date.isoformat(),
        INTERNAL_UNSET_FIELDS_FIELD: (
            "sections",
            "company",
            "participant",
            "portfolio",
            "attention",
            "contact",
            "country",
            "security_requirement_usd",
            *(
                market
                for market in ("fpm_m", "fpm_w", "dam", "idm")
                if market not in active_market_sections
            ),
        ),
        "currency": get_label_value(summary_sheet, "Currency:"),
        "market_turnover_usd": to_float(get_label_value(summary_sheet, "Market Turnover:")),
        "confirmed_trade_type": confirmed_trade_type,
        "balancing_market": balancing_market,
        "total_purchases": total_purchases,
        "total_sales": total_sales,
        "net_amount_traded_usd": net_amount_traded,
        "admin_fee_mwh": admin_fee["mwh"],
        "admin_fee_usd": admin_fee["amount_usd"],
        "wheeling_fee_usd": wheeling_fee,
        "losses_fee_usd": losses_fee,
        "total_fees_usd": total_fees,
        "total_amount_due_usd": total_amount_due,
        "gross_total_mwh": gross_total_mwh,
        "gross_total_amount_usd": gross_total_amount,
        "gross_average_price_usd_per_mwh": safe_divide(
            gross_total_amount,
            gross_total_mwh,
        ),
        "total_expenditure_usd": total_expenditure,
        "sapp_net_turnover_usd": total_fees,
        "metadata": {
            "data_source": data_source,
            "source_file": file_path.name,
        },
        "source_file": file_path.name,
    }
    for market, section in active_market_sections.items():
        record[market] = section

    return [
        record,
        *extract_trading_invoice_hourly_details(workbook, delivery_date, file_path),
    ]


def extract_trading_invoice_job(file_path: Path, job: SappExtractionJob) -> list[dict]:
    return extract_trading_invoice_results(file_path, data_source=job.data_source)


def build_unique_filter(record: dict, unique_key_fields: tuple[str, ...]) -> dict:
    missing_fields = [field for field in unique_key_fields if field not in record]
    if missing_fields:
        raise RuntimeError(
            f"Record is missing unique key fields for import: {', '.join(missing_fields)}"
        )
    return {field: record[field] for field in unique_key_fields}


def store_records_in_database(records: list[dict], job: SappExtractionJob) -> dict:
    if not records:
        raise RuntimeError("No SAPP records were provided for database import.")

    from app.db.database import get_db

    db = get_db()
    now = datetime.now(timezone.utc)

    grouped_records = {}
    for record in records:
        record_to_store = dict(record)
        collection_name = record_to_store.pop(
            INTERNAL_COLLECTION_FIELD,
            job.collection_name,
        )
        unique_key_fields = record_to_store.pop(
            INTERNAL_UNIQUE_KEY_FIELDS_FIELD,
            job.unique_key_fields,
        )
        unset_fields = record_to_store.pop(INTERNAL_UNSET_FIELDS_FIELD, ())
        grouped_records.setdefault((collection_name, unique_key_fields), []).append(
            (record_to_store, unset_fields)
        )

    imported = 0
    updated = 0
    for (collection_name, unique_key_fields), collection_items in grouped_records.items():
        collection = db[collection_name]
        operations = []
        for record, unset_fields in collection_items:
            update_document = {
                "$set": {
                    **record,
                    "updated_at": now,
                },
                "$setOnInsert": {
                    "created_at": now,
                },
            }
            if unset_fields:
                update_document["$unset"] = {field: "" for field in unset_fields}

            operations.append(
                UpdateOne(
                    build_unique_filter(record, unique_key_fields),
                    update_document,
                    upsert=True,
                )
            )

        result = collection.bulk_write(operations, ordered=False)
        imported += result.upserted_count
        updated += result.matched_count

    stored_records = [
        {
            key: value
            for key, value in record.items()
            if key
            not in (
                INTERNAL_COLLECTION_FIELD,
                INTERNAL_UNIQUE_KEY_FIELDS_FIELD,
                INTERNAL_UNSET_FIELDS_FIELD,
            )
        }
        for record in records
    ]
    summary_record = stored_records[0]
    return {
        "job": job.name,
        "delivery_date": summary_record["delivery_date"],
        "imported": imported,
        "updated": updated,
        "source_file": summary_record.get("source_file", ""),
    }


def store_results_in_database(records: list[dict]) -> dict:
    return store_records_in_database(records, CONSTRAINED_AREA_RESULTS_JOB)


def download_attachment(driver, download_dir: Path, extension: str = ".xlsx"):
    # Click the attachment button for the Excel file using the data-cy/title attribute.
    attachment_xpath = (
        f"//button[contains(@data-cy, {xpath_literal(extension)}) "
        f"or contains(@title, {xpath_literal(extension)})]"
    )

    attachment_button = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((By.XPATH, attachment_xpath))
    )

    existing_files = {p.name for p in download_dir.glob(f"*{extension}")}
    driver.execute_script(
        "arguments[0].scrollIntoView({block: 'center', inline: 'center'});",
        attachment_button,
    )
    time.sleep(0.5)
    try:
        attachment_button.click()
    except (ElementClickInterceptedException, StaleElementReferenceException, WebDriverException):
        attachment_button = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, attachment_xpath))
        )
        driver.execute_script("arguments[0].click();", attachment_button)

    downloaded_file = wait_for_new_download_file(
        download_dir,
        existing_files,
        extension=extension,
        timeout=60,
    )
    print(f"📥 Downloaded file: {downloaded_file.name}")
    return downloaded_file


CONSTRAINED_AREA_RESULTS_JOB = SappExtractionJob(
    name="constrained_area_results",
    subject_template=CONSTRAINED_AREA_SUBJECT_TEMPLATE,
    data_source=CONSTRAINED_AREA_DATA_SOURCE,
    collection_name="sapp_constrained_area_results",
    unique_key_fields=("delivery_date", "hour"),
    extractor=extract_constrained_area_job,
)

UNCONSTRAINED_AREA_RESULTS_JOB = SappExtractionJob(
    name="unconstrained_area_results",
    subject_template=UNCONSTRAINED_AREA_SUBJECT_TEMPLATE,
    data_source=UNCONSTRAINED_AREA_DATA_SOURCE,
    collection_name="sapp_unconstrained_area_results",
    unique_key_fields=("delivery_date", "hour"),
    extractor=extract_unconstrained_area_job,
)

PARTICIPANT_PORTFOLIO_RESULTS_JOB = SappExtractionJob(
    name="participant_portfolio_results",
    subject_template=PARTICIPANT_PORTFOLIO_SUBJECT_TEMPLATE,
    data_source=PARTICIPANT_PORTFOLIO_DATA_SOURCE,
    collection_name="sapp_participant_portfolio_results",
    unique_key_fields=("market", "delivery_date", "hour"),
    extractor=extract_participant_portfolio_job,
)

PARTICIPANT_PORTFOLIO_FPM_W_RESULTS_JOB = SappExtractionJob(
    name="participant_portfolio_results_fpm_w",
    subject_template=PARTICIPANT_PORTFOLIO_FPM_W_SUBJECT_TEMPLATE,
    data_source=PARTICIPANT_PORTFOLIO_FPM_W_DATA_SOURCE,
    collection_name="sapp_participant_portfolio_results",
    unique_key_fields=("market", "delivery_date", "hour"),
    extractor=extract_participant_portfolio_job,
)

PARTICIPANT_PORTFOLIO_FPM_M_RESULTS_JOB = SappExtractionJob(
    name="participant_portfolio_results_fpm_m",
    subject_template=PARTICIPANT_PORTFOLIO_FPM_M_SUBJECT_TEMPLATE,
    data_source=PARTICIPANT_PORTFOLIO_FPM_M_DATA_SOURCE,
    collection_name="sapp_participant_portfolio_results",
    unique_key_fields=("market", "delivery_date", "hour"),
    extractor=extract_participant_portfolio_job,
)

TRADING_INVOICE_RESULTS_JOB = SappExtractionJob(
    name="trading_invoice_credit_note",
    subject_template=TRADING_INVOICE_SUBJECT_TEMPLATE,
    data_source=TRADING_INVOICE_DATA_SOURCE,
    collection_name="sapp_trading_invoice_credit_notes",
    unique_key_fields=("delivery_date",),
    extractor=extract_trading_invoice_job,
)

SAPP_EXTRACTION_JOBS: dict[str, SappExtractionJob] = {
    CONSTRAINED_AREA_RESULTS_JOB.name: CONSTRAINED_AREA_RESULTS_JOB,
    UNCONSTRAINED_AREA_RESULTS_JOB.name: UNCONSTRAINED_AREA_RESULTS_JOB,
    PARTICIPANT_PORTFOLIO_RESULTS_JOB.name: PARTICIPANT_PORTFOLIO_RESULTS_JOB,
    PARTICIPANT_PORTFOLIO_FPM_W_RESULTS_JOB.name: PARTICIPANT_PORTFOLIO_FPM_W_RESULTS_JOB,
    PARTICIPANT_PORTFOLIO_FPM_M_RESULTS_JOB.name: PARTICIPANT_PORTFOLIO_FPM_M_RESULTS_JOB,
    TRADING_INVOICE_RESULTS_JOB.name: TRADING_INVOICE_RESULTS_JOB,
}

PARTICIPANT_PORTFOLIO_BUNDLE_JOB_NAMES = [
    "participant_portfolio_results",
    "participant_portfolio_results_fpm_w",
    "participant_portfolio_results_fpm_m",
]


def get_extraction_job(job_name: str) -> SappExtractionJob:
    try:
        return SAPP_EXTRACTION_JOBS[job_name]
    except KeyError as exc:
        available_jobs = ", ".join(sorted(SAPP_EXTRACTION_JOBS))
        raise ValueError(
            f"Unknown SAPP extraction job '{job_name}'. Available jobs: {available_jobs}"
        ) from exc


def iter_delivery_dates(start_date: date, end_date: date):
    if end_date < start_date:
        raise ValueError("end_date must be greater than or equal to start_date.")

    current_date = start_date
    while current_date <= end_date:
        yield current_date
        current_date += timedelta(days=1)


def iter_delivery_dates_descending(start_date: date, end_date: date):
    if end_date < start_date:
        raise ValueError("end_date must be greater than or equal to start_date.")

    current_date = end_date
    while current_date >= start_date:
        yield current_date
        current_date -= timedelta(days=1)


def run_extraction_job(
    job: SappExtractionJob,
    delivery_date: Optional[date] = None,
    page_start: int = 1,
) -> dict:
    delivery_date = delivery_date or datetime.now().date()
    if page_start < 1:
        raise ValueError("page_start must be greater than or equal to 1.")
    target_subject = job.build_subject(delivery_date)
    print(
        f"Starting SAPP scraper job '{job.name}' for {delivery_date.isoformat()} "
        f"from inbox page {page_start}"
    )
    username, password = load_config()
    with tempfile.TemporaryDirectory(prefix="sapp-download-") as temp_download_dir:
        download_dir = Path(temp_download_dir)
        driver = create_driver(download_dir)
        try:
            login(driver, username, password)
            navigate_to_inbox(driver)
            find_message_and_open(
                driver,
                job,
                delivery_date,
                page_start=page_start,
            )
            downloaded_file = download_attachment(
                driver,
                download_dir,
                extension=job.attachment_extension,
            )
            records = job.extractor(downloaded_file, job)
            result = store_records_in_database(records, job)
            result["page_start"] = page_start
            print(f"SAPP scraper completed successfully: {result}")
            return result
        finally:
            driver.quit()


def run_extraction_job_for_date_range(
    job: SappExtractionJob,
    start_date: date,
    end_date: date,
    continue_on_error: bool = True,
    page_start: int = 1,
) -> dict:
    if page_start < 1:
        raise ValueError("page_start must be greater than or equal to 1.")
    delivery_dates = list(iter_delivery_dates_descending(start_date, end_date))
    pending_dates = set(delivery_dates)
    print(
        f"Starting SAPP scraper job '{job.name}' for date range "
        f"{start_date.isoformat()} to {end_date.isoformat()} "
        f"from newest to oldest, starting at inbox page {page_start}"
    )

    username, password = load_config()
    results = []
    stop_reason = None

    with tempfile.TemporaryDirectory(prefix="sapp-download-") as temp_download_dir:
        download_dir = Path(temp_download_dir)
        driver = create_driver(download_dir)
        try:
            login(driver, username, password)
            navigate_to_inbox(driver)
            go_to_inbox_page(driver, page_start)

            for page_number in range(page_start, page_start + MAX_INBOX_PAGES_TO_SEARCH):
                if not pending_dates:
                    break

                wait_for_inbox_grid_to_settle(driver, timeout=INBOX_GRID_READY_TIMEOUT)
                print(
                    f"[5/7] Scanning inbox page {page_number} for "
                    f"{len(pending_dates)} remaining target messages"
                )

                while pending_dates:
                    found_subjects = find_matching_subjects_on_current_page(
                        driver,
                        job,
                        sorted(pending_dates, reverse=True),
                    )
                    if not found_subjects:
                        break

                    for target_subject, delivery_date in found_subjects.items():
                        if delivery_date not in pending_dates:
                            continue
                        try:
                            print(
                                f"[5/7] Found target message for "
                                f"{delivery_date.isoformat()} on inbox page {page_number}"
                            )
                            click_message(driver, target_subject)
                            print(
                                f"[5/7] Opened target message for "
                                f"{delivery_date.isoformat()} on inbox page {page_number}"
                            )
                            downloaded_file = download_attachment(
                                driver,
                                download_dir,
                                extension=job.attachment_extension,
                            )
                            records = job.extractor(downloaded_file, job)
                            result = store_records_in_database(records, job)
                            results.append(
                                {
                                    "delivery_date": delivery_date.isoformat(),
                                    "status": "success",
                                    **result,
                                }
                            )
                            pending_dates.discard(delivery_date)
                        except Exception as exc:
                            failure = {
                                "job": job.name,
                                "delivery_date": delivery_date.isoformat(),
                                "status": "failed",
                                "error": str(exc),
                            }
                            results.append(failure)
                            pending_dates.discard(delivery_date)
                            print(
                                f"SAPP scraper failed for "
                                f"{delivery_date.isoformat()}: {exc}"
                            )
                            if not continue_on_error:
                                raise

                if not pending_dates:
                    break

                oldest_pending_date = min(pending_dates)
                visible_job_dates = extract_job_delivery_dates_from_current_page(driver, job)
                if visible_job_dates and max(visible_job_dates) < oldest_pending_date:
                    stop_reason = (
                        "Visible matching inbox subjects are older than the oldest pending "
                        f"requested date {oldest_pending_date.isoformat()}"
                    )
                    print(
                        f"[5/7] Visible {job.name} subjects are older than the oldest "
                        f"pending date {oldest_pending_date.isoformat()}. "
                        "Stopping further page scans."
                    )
                    break

                next_page_button = find_enabled_next_page_button(driver)
                if next_page_button is None:
                    break

                print(
                    f"[5/7] Moving to inbox page {page_number + 1}; "
                    f"{len(pending_dates)} target messages still pending"
                )
                click_next_inbox_page(driver, page_number)

            for delivery_date in sorted(pending_dates, reverse=True):
                failure = {
                    "job": job.name,
                    "delivery_date": delivery_date.isoformat(),
                    "status": "failed",
                    "error": stop_reason
                    or (
                        "Target message not found after searching up to "
                        f"{MAX_INBOX_PAGES_TO_SEARCH} inbox pages from page "
                        f"{page_start}: {job.build_subject(delivery_date)}"
                    ),
                }
                results.append(failure)
                print(f"SAPP scraper did not find message for {delivery_date.isoformat()}")
                if not continue_on_error:
                    raise RuntimeError(failure["error"])
        finally:
            driver.quit()

    successful_results = [result for result in results if result["status"] == "success"]
    failed_results = [result for result in results if result["status"] == "failed"]
    return {
        "job": job.name,
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "page_start": page_start,
        "requested_dates": len(delivery_dates),
        "successful_dates": len(successful_results),
        "failed_dates": len(failed_results),
        "imported": sum(result.get("imported", 0) for result in successful_results),
        "updated": sum(result.get("updated", 0) for result in successful_results),
        "results": results,
    }


def run_portfolio_extraction_bundle(
    delivery_date: Optional[date] = None,
    page_start: int = 1,
) -> dict:
    delivery_date = delivery_date or datetime.now().date()
    if page_start < 1:
        raise ValueError("page_start must be greater than or equal to 1.")

    jobs = [get_extraction_job(job_name) for job_name in PARTICIPANT_PORTFOLIO_BUNDLE_JOB_NAMES]
    print(
        f"Starting bundled SAPP participant portfolio scrape for {delivery_date.isoformat()} "
        f"from inbox page {page_start}"
    )

    username, password = load_config()
    results = []
    pending_jobs = list(jobs)

    with tempfile.TemporaryDirectory(prefix="sapp-download-") as temp_download_dir:
        download_dir = Path(temp_download_dir)
        driver = create_driver(download_dir)
        try:
            login(driver, username, password)
            navigate_to_inbox(driver)
            go_to_inbox_page(driver, page_start)

            for page_number in range(page_start, page_start + MAX_INBOX_PAGES_TO_SEARCH):
                if not pending_jobs:
                    break

                wait_for_inbox_grid_to_settle(driver, timeout=INBOX_GRID_READY_TIMEOUT)
                print(
                    f"[5/7] Scanning inbox page {page_number} for "
                    f"{len(pending_jobs)} remaining participant portfolio subjects"
                )

                found_subjects = find_matching_job_subjects_on_current_page(
                    driver,
                    jobs,
                    [(job, delivery_date) for job in pending_jobs],
                )

                for target_subject, (job, matched_date) in found_subjects.items():
                    if job not in pending_jobs:
                        continue
                    try:
                        print(
                            f"[5/7] Found {job.name} message for "
                            f"{matched_date.isoformat()} on inbox page {page_number}"
                        )
                        click_message(driver, target_subject)
                        downloaded_file = download_attachment(
                            driver,
                            download_dir,
                            extension=job.attachment_extension,
                        )
                        records = job.extractor(downloaded_file, job)
                        result = store_records_in_database(records, job)
                        results.append(result)
                        pending_jobs.remove(job)
                    except Exception as exc:
                        results.append(
                            {
                                "job": job.name,
                                "delivery_date": matched_date.isoformat(),
                                "imported": 0,
                                "updated": 0,
                                "source_file": "",
                                "error": str(exc),
                                "status": "failed",
                            }
                        )
                        pending_jobs.remove(job)

                if not pending_jobs:
                    break

                next_page_button = find_enabled_next_page_button(driver)
                if next_page_button is None:
                    break
                click_next_inbox_page(driver, page_number)
        finally:
            driver.quit()

    successful_results = [
        result for result in results if result.get("error") in (None, "")
    ]
    failed_results = [result for result in results if result.get("error")]
    for job in pending_jobs:
        failed_results.append(
            {
                "job": job.name,
                "delivery_date": delivery_date.isoformat(),
                "imported": 0,
                "updated": 0,
                "source_file": "",
                "error": f"Target message not found after searching up to {MAX_INBOX_PAGES_TO_SEARCH} inbox pages from page {page_start}: {job.build_subject(delivery_date)}",
            }
        )

    all_results = successful_results + failed_results
    primary = successful_results[0] if successful_results else {
        "job": "participant_portfolio_results",
        "delivery_date": delivery_date.isoformat(),
        "imported": 0,
        "updated": 0,
        "source_file": "",
    }
    primary = dict(primary)
    primary.update(
        {
            "job": "participant_portfolio_results",
            "page_start": page_start,
            "requested_jobs": PARTICIPANT_PORTFOLIO_BUNDLE_JOB_NAMES,
            "successful_jobs": len(successful_results),
            "failed_jobs": len(failed_results),
            "imported": sum(result.get("imported", 0) for result in successful_results),
            "updated": sum(result.get("updated", 0) for result in successful_results),
            "related_results": all_results,
        }
    )
    return primary


def run_portfolio_extraction_bundle_for_date_range(
    start_date: date,
    end_date: date,
    continue_on_error: bool = True,
    page_start: int = 1,
) -> dict:
    if page_start < 1:
        raise ValueError("page_start must be greater than or equal to 1.")
    delivery_dates = list(iter_delivery_dates_descending(start_date, end_date))
    jobs = [get_extraction_job(job_name) for job_name in PARTICIPANT_PORTFOLIO_BUNDLE_JOB_NAMES]
    pending_items = {(job.name, delivery_date): job for job in jobs for delivery_date in delivery_dates}
    print(
        f"Starting bundled SAPP participant portfolio scrape for date range "
        f"{start_date.isoformat()} to {end_date.isoformat()} from newest to oldest, "
        f"starting at inbox page {page_start}"
    )

    username, password = load_config()
    results = []
    stop_reason = None

    with tempfile.TemporaryDirectory(prefix="sapp-download-") as temp_download_dir:
        download_dir = Path(temp_download_dir)
        driver = create_driver(download_dir)
        try:
            login(driver, username, password)
            navigate_to_inbox(driver)
            go_to_inbox_page(driver, page_start)

            for page_number in range(page_start, page_start + MAX_INBOX_PAGES_TO_SEARCH):
                if not pending_items:
                    break

                wait_for_inbox_grid_to_settle(driver, timeout=INBOX_GRID_READY_TIMEOUT)
                print(
                    f"[5/7] Scanning inbox page {page_number} for "
                    f"{len(pending_items)} remaining participant portfolio messages"
                )

                while pending_items:
                    ordered_pending = sorted(
                        ((job, delivery_date) for (job_name, delivery_date), job in pending_items.items()),
                        key=lambda item: item[1],
                        reverse=True,
                    )
                    found_subjects = find_matching_job_subjects_on_current_page(
                        driver,
                        jobs,
                        ordered_pending,
                    )
                    if not found_subjects:
                        break

                    for target_subject, (job, delivery_date) in found_subjects.items():
                        key = (job.name, delivery_date)
                        if key not in pending_items:
                            continue
                        try:
                            print(
                                f"[5/7] Found {job.name} message for "
                                f"{delivery_date.isoformat()} on inbox page {page_number}"
                            )
                            click_message(driver, target_subject)
                            downloaded_file = download_attachment(
                                driver,
                                download_dir,
                                extension=job.attachment_extension,
                            )
                            records = job.extractor(downloaded_file, job)
                            result = store_records_in_database(records, job)
                            results.append(
                                {
                                    "delivery_date": delivery_date.isoformat(),
                                    "status": "success",
                                    **result,
                                }
                            )
                            pending_items.pop(key, None)
                        except Exception as exc:
                            results.append(
                                {
                                    "job": job.name,
                                    "delivery_date": delivery_date.isoformat(),
                                    "status": "failed",
                                    "error": str(exc),
                                }
                            )
                            pending_items.pop(key, None)
                            if not continue_on_error:
                                raise

                if not pending_items:
                    break

                oldest_pending_date = min(delivery_date for _, delivery_date in pending_items.keys())
                visible_job_dates = sorted(
                    {
                        visible_date
                        for job in jobs
                        for visible_date in extract_job_delivery_dates_from_current_page(driver, job)
                    }
                )
                if visible_job_dates and max(visible_job_dates) < oldest_pending_date:
                    stop_reason = (
                        "Visible participant portfolio subjects are older than the oldest pending "
                        f"requested date {oldest_pending_date.isoformat()}"
                    )
                    print(f"[5/7] {stop_reason}. Stopping further page scans.")
                    break

                next_page_button = find_enabled_next_page_button(driver)
                if next_page_button is None:
                    break
                click_next_inbox_page(driver, page_number)

            for (job_name, delivery_date), job in sorted(
                pending_items.items(),
                key=lambda item: item[0][1],
                reverse=True,
            ):
                failure = {
                    "job": job_name,
                    "delivery_date": delivery_date.isoformat(),
                    "status": "failed",
                    "error": stop_reason
                    or (
                        "Target message not found after searching up to "
                        f"{MAX_INBOX_PAGES_TO_SEARCH} inbox pages from page {page_start}: "
                        f"{job.build_subject(delivery_date)}"
                    ),
                }
                results.append(failure)
                if not continue_on_error:
                    raise RuntimeError(failure["error"])
        finally:
            driver.quit()

    successful_results = [result for result in results if result["status"] == "success"]
    failed_results = [result for result in results if result["status"] == "failed"]
    return {
        "job": "participant_portfolio_results",
        "start_date": start_date.isoformat(),
        "end_date": end_date.isoformat(),
        "page_start": page_start,
        "requested_dates": len(delivery_dates) * len(jobs),
        "successful_dates": len(successful_results),
        "failed_dates": len(failed_results),
        "imported": sum(result.get("imported", 0) for result in successful_results),
        "updated": sum(result.get("updated", 0) for result in successful_results),
        "results": results,
        "requested_jobs": PARTICIPANT_PORTFOLIO_BUNDLE_JOB_NAMES,
        "successful_jobs": len([job for job in PARTICIPANT_PORTFOLIO_BUNDLE_JOB_NAMES if any(r.get("job") == job and r.get("status") == "success" for r in results)]),
        "failed_jobs": len([job for job in PARTICIPANT_PORTFOLIO_BUNDLE_JOB_NAMES if not any(r.get("job") == job and r.get("status") == "success" for r in results)]),
    }


def run_scraper(delivery_date: Optional[date] = None, page_start: int = 1) -> dict:
    return run_extraction_job(
        CONSTRAINED_AREA_RESULTS_JOB,
        delivery_date=delivery_date,
        page_start=page_start,
    )


def main():
    return run_scraper()


if __name__ == "__main__":
    main()
